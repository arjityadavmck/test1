#!/usr/bin/env python3
"""
Test Excel Quality Score Visibility Issue
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import tempfile

# Add the parent directory to sys.path
sys.path.insert(0, str(Path(__file__).parent))

from src.core import chat, parse_json_safely, score_test_cases

def test_excel_quality_scores():
    """Test Excel export with quality scores to debug visibility issue"""
    print("🔍 Testing Excel Quality Score Visibility")
    print("=" * 50)
    
    # Create sample test cases that should have quality scores
    test_cases = [
        {
            "id": "TC-001",
            "title": "User login with valid credentials",
            "steps": [
                "Navigate to the login page",
                "Enter valid username 'testuser@example.com'",
                "Enter valid password 'Password123!'",
                "Click the 'Login' button"
            ],
            "expected": "User is successfully logged in and redirected to the dashboard",
            "priority": "High"
        },
        {
            "id": "TC-002",
            "title": "User login with invalid password",
            "steps": [
                "Navigate to the login page",
                "Enter valid username 'testuser@example.com'",
                "Enter invalid password 'wrongpassword'",
                "Click the 'Login' button"
            ],
            "expected": "Error message 'Invalid email or password' is displayed",
            "priority": "High"
        },
        {
            "id": "TC-003",
            "title": "Password reset functionality",
            "steps": [
                "Navigate to the login page",
                "Click 'Forgot Password?' link",
                "Enter registered email address",
                "Click 'Send Reset Email' button"
            ],
            "expected": "Password reset email is sent to the user",
            "priority": "Medium"
        }
    ]
    
    # Create a sample requirement text
    requirement_text = """
    User Login System Requirements:
    
    The system must provide secure user authentication functionality.
    Users should be able to log in with email and password.
    Invalid login attempts should show appropriate error messages.
    Password reset functionality should be available for users who forget their passwords.
    """
    
    # Setup paths
    output_dir = Path("outputs/testcase_generated")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📋 Test Cases: {len(test_cases)}")
    for i, case in enumerate(test_cases, 1):
        print(f"   {i}. {case['id']}: {case['title']}")
    
    # Step 1: Generate quality scores
    print("\n📊 Step 1: Generating quality scores...")
    try:
        quality_report = score_test_cases(test_cases, requirement_text, output_dir)
        print(f"✅ Quality assessment completed")
        
        if quality_report:
            overall_score = quality_report.get("overall_score", 0)
            individual_scores = quality_report.get("individual_scores", [])
            print(f"   Overall Score: {overall_score:.1f}/10")
            print(f"   Individual Scores: {len(individual_scores)} test cases assessed")
            
            # Show individual scores
            for score_info in individual_scores:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                print(f"   - {test_id}: {total_score:.1f}/10")
        else:
            print("❌ No quality report generated")
            return False
            
    except Exception as e:
        print(f"❌ Quality assessment failed: {e}")
        return False
    
    # Step 2: Test Excel export with quality scores
    print("\n📊 Step 2: Testing Excel export...")
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"quality_test_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Headers
        headers = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Create quality score mapping
        quality_scores = {}
        if quality_report and "individual_scores" in quality_report:
            for score_info in quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                quality_scores[test_id] = total_score
                print(f"   Mapped {test_id} → {total_score:.1f}/10")
        
        print(f"   Quality scores mapping: {quality_scores}")
        
        # Data rows
        for row, case in enumerate(test_cases, 2):
            steps = case.get("steps", [])
            steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            print(f"   Row {row}: {test_id} → Quality: {quality_display}")
            
            # Add data to Excel
            ws.cell(row=row, column=1, value=test_id)
            ws.cell(row=row, column=2, value=case.get("title", ""))
            
            steps_cell = ws.cell(row=row, column=3, value=steps_text)
            steps_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.cell(row=row, column=4, value=case.get("expected", ""))
            ws.cell(row=row, column=5, value=case.get("priority", "Medium"))
            
            # This is the critical line - setting quality score
            quality_cell = ws.cell(row=row, column=6, value=quality_display)
            print(f"   Set cell F{row} = '{quality_display}'")
            
            # Color code quality scores
            if quality_score >= 8.0:
                quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                print(f"   Applied GREEN color to {test_id}")
            elif quality_score >= 6.0:
                quality_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
                print(f"   Applied YELLOW color to {test_id}")
            elif quality_score > 0:
                quality_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                print(f"   Applied PINK color to {test_id}")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        
        # Save Excel file
        wb.save(excel_file)
        print(f"✅ Excel file saved: {excel_file}")
        
        # Verify by reading back
        print("\n🔍 Step 3: Verifying Excel content...")
        wb_read = openpyxl.load_workbook(excel_file)
        ws_read = wb_read.active
        
        # Check headers
        headers_read = []
        for col in range(1, 7):
            header_value = ws_read.cell(row=1, column=col).value
            headers_read.append(header_value)
        print(f"   Headers: {headers_read}")
        
        # Check data rows
        for row in range(2, ws_read.max_row + 1):
            test_id = ws_read.cell(row=row, column=1).value
            quality_value = ws_read.cell(row=row, column=6).value
            print(f"   Row {row}: {test_id} → Quality Column: '{quality_value}'")
            
            if quality_value is None or quality_value == "N/A":
                print(f"   ❌ No quality score found for {test_id}")
            else:
                print(f"   ✅ Quality score found for {test_id}: {quality_value}")
        
        wb_read.close()
        
        print(f"\n📊 Excel Export Test Results:")
        print(f"   File: {excel_file}")
        print(f"   Size: {excel_file.stat().st_size} bytes")
        print(f"   Rows: {ws_read.max_row} (including header)")
        print(f"   Columns: {ws_read.max_column}")
        
        return True
        
    except ImportError:
        print("❌ openpyxl not available")
        return False
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Run the Excel quality score test"""
    try:
        success = test_excel_quality_scores()
        if success:
            print("\n✅ Excel quality score test completed!")
            print("\n📝 Next Steps:")
            print("1. Check the generated Excel file in outputs/testcase_generated/")
            print("2. Verify that quality scores are visible in column F")
            print("3. Check for color coding of quality scores")
        else:
            print("\n❌ Excel quality score test failed!")
            print("Check the errors above and ensure:")
            print("1. openpyxl is installed: pip install openpyxl")
            print("2. Quality assessment is working properly")
            print("3. Excel export logic is correct")
    except Exception as e:
        print(f"\n❌ Test failed with error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()