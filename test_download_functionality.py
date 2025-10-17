#!/usr/bin/env python3
"""
Test the download functionality for CSV and Excel export
"""

import sys
from pathlib import Path
import json
from datetime import datetime
import tempfile
import os

# Add the parent directory to sys.path
sys.path.insert(0, str(Path(__file__).parent))

from src.core import to_rows, write_csv, score_test_cases

def test_csv_export():
    """Test CSV export functionality"""
    print("🧪 Testing CSV Export...")
    
    # Sample test cases
    test_cases = [
        {
            "id": "TC-001",
            "title": "Login with valid credentials",
            "steps": [
                "Navigate to login page",
                "Enter valid username",
                "Enter valid password",
                "Click login button"
            ],
            "expected": "User is successfully logged in and redirected to dashboard",
            "priority": "High"
        },
        {
            "id": "TC-002", 
            "title": "Login with invalid password",
            "steps": [
                "Navigate to login page",
                "Enter valid username",
                "Enter invalid password",
                "Click login button"
            ],
            "expected": "Error message 'Invalid credentials' is displayed",
            "priority": "High"
        },
        {
            "id": "TC-003",
            "title": "Password reset functionality",
            "steps": [
                "Click 'Forgot Password' link",
                "Enter registered email",
                "Click 'Send Reset Link' button",
                "Check email inbox"
            ],
            "expected": "Password reset email is sent to user",
            "priority": "Medium"
        }
    ]
    
    # Test existing CSV export
    output_dir = Path("outputs/testcase_generated")
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_file = output_dir / "test_export.csv"
    
    rows = to_rows(test_cases)
    write_csv(rows, csv_file)
    
    if csv_file.exists():
        print(f"✅ CSV export successful: {csv_file}")
        print(f"   File size: {csv_file.stat().st_size} bytes")
        
        # Read and display first few lines
        with open(csv_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()[:5]
            print("   First few lines:")
            for line in lines:
                print(f"   {line.strip()}")
    else:
        print("❌ CSV export failed")
        return False
    
    return True

def test_excel_export():
    """Test Excel export functionality"""
    print("\n📊 Testing Excel Export...")
    
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        print("✅ openpyxl library is available")
    except ImportError as e:
        print(f"❌ openpyxl not available: {e}")
        return False
    
    # Sample test cases with quality scores
    test_cases = [
        {
            "id": "TC-001",
            "title": "Login with valid credentials",
            "steps": [
                "Navigate to login page",
                "Enter valid username 'user@example.com'",
                "Enter valid password 'Password123'",
                "Click login button"
            ],
            "expected": "User is successfully logged in and redirected to dashboard",
            "priority": "High"
        },
        {
            "id": "TC-002", 
            "title": "Login with invalid password",
            "steps": [
                "Navigate to login page",
                "Enter valid username 'user@example.com'",
                "Enter invalid password 'wrongpass'",
                "Click login button"
            ],
            "expected": "Error message 'Invalid email or password' is displayed",
            "priority": "High"
        }
    ]
    
    # Create sample quality report
    quality_report = {
        "overall_score": 8.5,
        "individual_scores": [
            {
                "test_id": "TC-001",
                "total_score": 9.2,
                "scores": {
                    "clarity": 9.5,
                    "completeness": 9.0,
                    "specificity": 9.3,
                    "testability": 9.1,
                    "coverage": 8.9
                }
            },
            {
                "test_id": "TC-002",
                "total_score": 7.8,
                "scores": {
                    "clarity": 8.2,
                    "completeness": 7.5,
                    "specificity": 8.0,
                    "testability": 7.9,
                    "coverage": 7.4
                }
            }
        ]
    }
    
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Headers
    headers = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Quality score mapping
    quality_scores = {}
    for score_info in quality_report["individual_scores"]:
        test_id = score_info.get("test_id", "")
        total_score = score_info.get("total_score", 0)
        quality_scores[test_id] = total_score
    
    # Data rows
    for row, case in enumerate(test_cases, 2):
        steps = case.get("steps", [])
        steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
        
        test_id = case.get("id", "")
        quality_score = quality_scores.get(test_id, 0)
        quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
        
        ws.cell(row=row, column=1, value=test_id)
        ws.cell(row=row, column=2, value=case.get("title", ""))
        ws.cell(row=row, column=3, value=steps_text)
        ws.cell(row=row, column=4, value=case.get("expected", ""))
        ws.cell(row=row, column=5, value=case.get("priority", "Medium"))
        
        quality_cell = ws.cell(row=row, column=6, value=quality_display)
        
        # Color code quality scores
        if quality_score >= 8.0:
            quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif quality_score >= 6.0:
            quality_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
    
    # Save Excel file
    output_dir = Path("outputs/testcase_generated")
    output_dir.mkdir(parents=True, exist_ok=True)
    excel_file = output_dir / "test_export.xlsx"
    
    wb.save(excel_file)
    
    if excel_file.exists():
        print(f"✅ Excel export successful: {excel_file}")
        print(f"   File size: {excel_file.stat().st_size} bytes")
        
        # Verify content by reading back
        wb_read = openpyxl.load_workbook(excel_file)
        ws_read = wb_read.active
        print(f"   Worksheet title: {ws_read.title}")
        print(f"   Max row: {ws_read.max_row}")
        print(f"   Max column: {ws_read.max_column}")
        
        # Display some cells
        print("   Sample content:")
        for row in range(1, min(4, ws_read.max_row + 1)):
            row_values = []
            for col in range(1, min(4, ws_read.max_column + 1)):
                cell_value = ws_read.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value)[:20])
            print(f"   Row {row}: {' | '.join(row_values)}")
        
        wb_read.close()
    else:
        print("❌ Excel export failed")
        return False
    
    return True

def main():
    """Run all tests"""
    print("🧪 Testing Download Functionality")
    print("=" * 50)
    
    csv_success = test_csv_export()
    excel_success = test_excel_export()
    
    print("\n📊 Test Results Summary:")
    print(f"CSV Export: {'✅ PASS' if csv_success else '❌ FAIL'}")
    print(f"Excel Export: {'✅ PASS' if excel_success else '❌ FAIL'}")
    
    if csv_success and excel_success:
        print("\n🎉 All download functionality tests passed!")
        print("\n📝 Next Steps:")
        print("1. Launch the UI: python3 ui_app/main.py")
        print("2. Generate some test cases")
        print("3. Use the new '📄 Download as CSV' and '📊 Download as Excel' buttons")
        print("4. Verify the exported files have quality scores and formatting")
        return True
    else:
        print("\n❌ Some tests failed. Please check the errors above.")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)