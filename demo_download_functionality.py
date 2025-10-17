#!/usr/bin/env python3
"""
Demo Script: Generate Test Cases with Quality Scores and Download Options
=========================================================================

This script demonstrates the enhanced TestCase Generator functionality including:
1. Requirement enhancement
2. Test case generation with quality scoring
3. Export to CSV and Excel formats
"""

import sys
from pathlib import Path
import json
from datetime import datetime

# Add the parent directory to sys.path
sys.path.insert(0, str(Path(__file__).parent))

from src.core import chat, parse_json_safely, to_rows, write_csv, score_test_cases, enhance_requirement

def demo_complete_workflow():
    """Demonstrate the complete enhanced workflow"""
    print("🚀 TestCase Generator - Complete Workflow Demo")
    print("=" * 60)
    
    # Setup paths
    ROOT = Path(__file__).resolve().parent
    REQ_DIR = ROOT / "data" / "requirements"
    OUT_DIR = ROOT / "outputs" / "testcase_generated"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    
    PROMPTS_DIR = ROOT / "src" / "core" / "prompts"
    SYSTEM_PROMPT = (PROMPTS_DIR / "testcase_system.txt").read_text(encoding="utf-8")
    USER_PROMPT_TEMPLATE = (PROMPTS_DIR / "testcase_user.txt").read_text(encoding="utf-8")
    
    # Step 1: Load a requirement file
    print("\n📄 Step 1: Loading requirement file...")
    req_file = REQ_DIR / "login.txt"
    if not req_file.exists():
        print(f"❌ Requirement file not found: {req_file}")
        return False
    
    original_requirement = req_file.read_text(encoding="utf-8").strip()
    print(f"✅ Loaded requirement: {req_file.name}")
    print(f"   Original length: {len(original_requirement)} characters")
    
    # Step 2: Enhance requirement
    print("\n📝 Step 2: Enhancing requirement for better test generation...")
    try:
        enhanced_requirement, enhancement_report = enhance_requirement(original_requirement, OUT_DIR)
        print(f"✅ Requirement enhanced successfully!")
        print(f"   Enhanced length: {len(enhanced_requirement)} characters")
        print(f"   Quality score: {enhancement_report.get('overall_score', 0):.1f}/10")
        print(f"   Improvements made: {len(enhancement_report.get('improvements_made', []))}")
    except Exception as e:
        print(f"⚠️ Enhancement failed: {e}")
        print("   Proceeding with original requirement...")
        enhanced_requirement = original_requirement
        enhancement_report = {}
    
    # Step 3: Generate test cases
    print("\n🤖 Step 3: Generating test cases...")
    try:
        user_prompt = USER_PROMPT_TEMPLATE.format(requirement_text=enhanced_requirement)
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt}
        ]
        
        raw_response = chat(messages)
        test_cases = parse_json_safely(raw_response, OUT_DIR / "demo_raw.json")
        
        if not test_cases:
            print("❌ No test cases generated")
            return False
        
        print(f"✅ Generated {len(test_cases)} test cases")
        
        # Display test case summaries
        for i, case in enumerate(test_cases[:3], 1):  # Show first 3
            print(f"   {i}. {case.get('id', 'N/A')}: {case.get('title', 'No title')}")
    
    except Exception as e:
        print(f"❌ Test case generation failed: {e}")
        return False
    
    # Step 4: Quality assessment
    print("\n📊 Step 4: Assessing test case quality...")
    try:
        quality_report = score_test_cases(test_cases, enhanced_requirement, OUT_DIR)
        overall_score = quality_report.get("overall_score", 0)
        individual_scores = quality_report.get("individual_scores", [])
        
        print(f"✅ Quality assessment completed!")
        print(f"   Overall quality score: {overall_score:.1f}/10")
        
        # Show quality distribution
        high_quality = sum(1 for score in individual_scores if score.get("total_score", 0) >= 8.0)
        medium_quality = sum(1 for score in individual_scores if 6.0 <= score.get("total_score", 0) < 8.0)
        low_quality = sum(1 for score in individual_scores if score.get("total_score", 0) < 6.0)
        
        print(f"   Quality distribution: 🟢{high_quality} High | 🟡{medium_quality} Medium | 🔴{low_quality} Low")
        
    except Exception as e:
        print(f"⚠️ Quality assessment failed: {e}")
        quality_report = {}
    
    # Step 5: Export to CSV
    print("\n📄 Step 5: Generating CSV export...")
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_file = OUT_DIR / f"demo_export_{timestamp}.csv"
        
        # Create CSV with quality scores
        import csv
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            # Create quality score mapping
            quality_scores = {}
            if quality_report and "individual_scores" in quality_report:
                for score_info in quality_report["individual_scores"]:
                    test_id = score_info.get("test_id", "")
                    total_score = score_info.get("total_score", 0)
                    quality_scores[test_id] = total_score
            
            for case in test_cases:
                steps = case.get("steps", [])
                if isinstance(steps, list):
                    steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
                else:
                    steps_text = str(steps)
                
                test_id = case.get("id", "")
                quality_score = quality_scores.get(test_id, 0)
                quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
                
                writer.writerow({
                    'Test ID': test_id,
                    'Title': case.get("title", ""),
                    'Steps': steps_text,
                    'Expected Result': case.get("expected", ""),
                    'Priority': case.get("priority", "Medium"),
                    'Quality Score': quality_display
                })
        
        print(f"✅ CSV export successful: {csv_file}")
        print(f"   File size: {csv_file.stat().st_size} bytes")
        
    except Exception as e:
        print(f"❌ CSV export failed: {e}")
    
    # Step 6: Export to Excel
    print("\n📊 Step 6: Generating Excel export...")
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        excel_file = OUT_DIR / f"demo_export_{timestamp}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Headers
        headers = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows with quality scores and formatting
        for row, case in enumerate(test_cases, 2):
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            ws.cell(row=row, column=1, value=test_id)
            ws.cell(row=row, column=2, value=case.get("title", ""))
            
            steps_cell = ws.cell(row=row, column=3, value=steps_text)
            steps_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws.cell(row=row, column=4, value=case.get("expected", ""))
            
            priority_cell = ws.cell(row=row, column=5, value=case.get("priority", "Medium"))
            quality_cell = ws.cell(row=row, column=6, value=quality_display)
            
            # Color code quality scores
            if quality_score >= 8.0:
                quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif quality_score >= 6.0:
                quality_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
            elif quality_score > 0:
                quality_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        
        wb.save(excel_file)
        
        print(f"✅ Excel export successful: {excel_file}")
        print(f"   File size: {excel_file.stat().st_size} bytes")
        
    except ImportError:
        print("⚠️ Excel export skipped (openpyxl not available)")
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
    
    # Summary
    print("\n🎉 Complete Workflow Demo Summary:")
    print("=" * 40)
    print(f"✅ Requirement enhanced: {len(enhancement_report.get('improvements_made', []))} improvements")
    print(f"✅ Test cases generated: {len(test_cases)} cases")
    print(f"✅ Quality assessment: {overall_score:.1f}/10 overall score")
    print(f"✅ CSV export: Available in outputs/")
    print(f"✅ Excel export: Available in outputs/")
    
    print("\n📋 Generated Files:")
    output_files = list(OUT_DIR.glob(f"demo_*{timestamp}*"))
    for file in output_files:
        print(f"   📄 {file.name} ({file.stat().st_size} bytes)")
    
    print("\n🚀 Next Steps:")
    print("1. Launch the desktop UI: python3 ui_app/main.py")
    print("2. Generate test cases and click the download buttons")
    print("3. Test the CSV and Excel export functionality")
    print("4. Review the generated files with quality scores and formatting")
    
    return True

def main():
    """Run the complete workflow demo"""
    try:
        success = demo_complete_workflow()
        if success:
            print("\n✅ Demo completed successfully!")
        else:
            print("\n❌ Demo encountered some issues.")
    except KeyboardInterrupt:
        print("\n\n🛑 Demo interrupted by user")
    except Exception as e:
        print(f"\n❌ Demo failed with error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()