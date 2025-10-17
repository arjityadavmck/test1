#!/usr/bin/env python3
"""
TestCase Generator - Desktop UI Application
==========================================

A GUI wrapper around the existing testcase agent functionality.
Built with Tkinter for cross-platform compatibility.

Run: python ui_app/main.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import sys
from pathlib import Path
import sv_ttk
import logging
import csv
import json
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Add the parent directory to sys.path so we can import from src
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.core import chat, pick_requirement, parse_json_safely, to_rows, write_csv, score_test_cases, enhance_requirement
from src.integrations.testrail import map_case_to_testrail_payload, create_case, list_cases, add_result, get_stats


class TestCaseGeneratorApp:
    """Main application window for the TestCase Generator."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("M+ TGen")
        self.root.geometry("1450x950")
        self.root.minsize(1200, 700)
        
        # Configure style
        self.style = ttk.Style()
        self.style.theme_use('clam')  # Use a modern theme
        
        # Initialize variables
        self.requirement_path = tk.StringVar()
        self.requirement_text = ""
        self.original_requirement_text = ""
        self.enhancement_report = {}
        self.generated_cases = []
        self.quality_report = {}
        self.is_processing = False
        self.is_placeholder_text = True
        
        # Setup paths (reuse from existing agent)
        self.ROOT = Path(__file__).resolve().parents[1]
        self.REQ_DIR = self.ROOT / "data" / "requirements"
        self.OUT_DIR = self.ROOT / "outputs" / "testcase_generated"
        self.OUT_DIR.mkdir(parents=True, exist_ok=True)
        self.OUT_CSV = self.OUT_DIR / "test_cases.csv"
        self.LAST_RAW_JSON = self.OUT_DIR / "last_raw.json"
        
        # Load prompts
        self.PROMPTS_DIR = self.ROOT / "src" / "core" / "prompts"
        self.SYSTEM_PROMPT = (self.PROMPTS_DIR / "testcase_system.txt").read_text(encoding="utf-8")
        self.USER_PROMPT_TEMPLATE = (self.PROMPTS_DIR / "testcase_user.txt").read_text(encoding="utf-8")
        
        # Setup logging
        self.setup_logging()
        
        # Build UI
        self.create_widgets()
        
        # Initialize test type description
        self.on_test_type_change()
        
        # Load default requirement if available
        self.load_default_requirement()
    
    def setup_logging(self):
        """Configure logging to display in the UI."""
        self.log_handler = LogHandler(self)
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[self.log_handler]
        )
        self.logger = logging.getLogger(__name__)
    
    def create_scrollable_frame(self):
        """Create a scrollable frame for the entire application."""
        # Create main canvas and scrollbar
        self.canvas = tk.Canvas(self.root, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas, padding="20")
        
        # Configure scrolling
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        # Create window in canvas
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Configure canvas scrolling
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Bind mouse wheel scrolling
        self.bind_mousewheel()
        
        # Configure canvas to expand with window
        self.canvas.bind('<Configure>', self.on_canvas_configure)
        
        # Add keyboard scrolling support
        self.canvas.bind('<Key>', self.on_key_scroll)
        self.canvas.focus_set()
        
        # Grid the canvas and scrollbar
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.scrollbar.grid(row=0, column=1, sticky="ns")
    
    def bind_mousewheel(self):
        """Bind mouse wheel events for scrolling."""
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _on_mousewheel_linux(event):
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")
        
        # Bind scrolling to the canvas and the entire application
        widgets_to_bind = [self.canvas, self.root, self.scrollable_frame]
        
        for widget in widgets_to_bind:
            if sys.platform == "darwin":  # macOS
                widget.bind("<MouseWheel>", _on_mousewheel)
            elif sys.platform == "win32":  # Windows  
                widget.bind("<MouseWheel>", _on_mousewheel)
            else:  # Linux
                widget.bind("<Button-4>", _on_mousewheel_linux)
                widget.bind("<Button-5>", _on_mousewheel_linux)
    
    def on_canvas_configure(self, event):
        """Configure canvas scrolling when window is resized."""
        # Update the scrollable frame width to match canvas width
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)
    
    def on_key_scroll(self, event):
        """Handle keyboard scrolling."""
        if event.keysym == 'Up':
            self.canvas.yview_scroll(-1, "units")
        elif event.keysym == 'Down':
            self.canvas.yview_scroll(1, "units")
        elif event.keysym == 'Prior':  # Page Up
            self.canvas.yview_scroll(-5, "units")
        elif event.keysym == 'Next':  # Page Down
            self.canvas.yview_scroll(5, "units")
        elif event.keysym == 'Home':
            self.canvas.yview_moveto(0)
        elif event.keysym == 'End':
            self.canvas.yview_moveto(1)
    
    def bind_mousewheel_to_widget(self, widget):
        """Bind mouse wheel scrolling to a specific widget."""
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _on_mousewheel_linux(event):
            if event.num == 4:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.canvas.yview_scroll(1, "units")
        
        if sys.platform == "darwin":  # macOS
            widget.bind("<MouseWheel>", _on_mousewheel)
        elif sys.platform == "win32":  # Windows
            widget.bind("<MouseWheel>", _on_mousewheel)
        else:  # Linux
            widget.bind("<Button-4>", _on_mousewheel_linux)
            widget.bind("<Button-5>", _on_mousewheel_linux)
    
    def create_widgets(self):
        """Create and layout all UI widgets."""
        # Configure modern styling first
        style = ttk.Style()
        
        # Set consistent fonts for better typography
        try:
            # Use system fonts for better cross-platform appearance
            if sys.platform == "darwin":  # macOS
                default_font = ("SF Pro Display", 10)
                heading_font = ("SF Pro Display", 12, "bold") 
            elif sys.platform == "win32":  # Windows
                default_font = ("Segoe UI", 9)
                heading_font = ("Segoe UI", 11, "bold")
            else:  # Linux
                default_font = ("Cantarell", 10)
                heading_font = ("Cantarell", 12, "bold")
            
            style.configure("TLabel", font=default_font)
            style.configure("TButton", font=default_font)
            style.configure("Heading.TLabel", font=heading_font)
        except Exception:
            pass  # Fall back to default fonts if system fonts unavailable
            
        # Improve Treeview appearance with better row height and spacing
        style.configure("Treeview", rowheight=45, font=("Consolas", 10))
        style.configure("Treeview.Heading", font=("Consolas", 10, "bold"))
        
        # Add some padding inside treeview cells
        style.map("Treeview", fieldbackground=[('selected', '#0078d4')])
        style.configure("Treeview", foreground="black", background="white")
        
        # Create scrollable main container
        self.create_scrollable_frame()
        
        # Configure root grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Configure main frame grid weights
        main_frame = self.scrollable_frame
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        main_frame.rowconfigure(4, weight=1)
        main_frame.rowconfigure(7, weight=1)
        
        # Title
        # title_label = ttk.Label(main_frame, text="🤖 TestCase Generator", 
        #                        font=('Arial', 16, 'bold'))
        # title_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))
        
        # Enhanced Requirement Input section with better spacing
        req_input_frame = ttk.LabelFrame(main_frame, text="📄 Requirement Input", padding="20")
        req_input_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 20))
        req_input_frame.columnconfigure(0, weight=1)
        req_input_frame.rowconfigure(4, weight=1)
        
        # Input method selection
        input_method_frame = ttk.Frame(req_input_frame)
        input_method_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        input_method_frame.columnconfigure(1, weight=1)
        
        self.input_method = tk.StringVar(value="manual")
        
        manual_rb = ttk.Radiobutton(input_method_frame, text="✏️ Type requirements manually", 
                                   variable=self.input_method, value="manual",
                                   command=self.on_input_method_change)
        manual_rb.grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        
        file_rb = ttk.Radiobutton(input_method_frame, text="📁 Load from file", 
                                 variable=self.input_method, value="file",
                                 command=self.on_input_method_change)
        file_rb.grid(row=0, column=1, sticky=tk.W)
        
        # Test type selection
        test_type_frame = ttk.LabelFrame(req_input_frame, text="🎯 Test Type Selection", padding="10")
        test_type_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(10, 15))
        test_type_frame.columnconfigure(2, weight=1)
        
        self.test_type = tk.StringVar(value="smoke")
        
        smoke_rb = ttk.Radiobutton(test_type_frame, text="Smoke Tests", 
                                  variable=self.test_type, value="smoke",
                                  command=self.on_test_type_change)
        smoke_rb.grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        
        sanity_rb = ttk.Radiobutton(test_type_frame, text="Sanity Tests", 
                                   variable=self.test_type, value="sanity",
                                   command=self.on_test_type_change)
        sanity_rb.grid(row=0, column=1, sticky=tk.W, padx=(0, 20))
        
        unit_rb = ttk.Radiobutton(test_type_frame, text="Unit Tests", 
                                 variable=self.test_type, value="unit",
                                 command=self.on_test_type_change)
        unit_rb.grid(row=0, column=2, sticky=tk.W, padx=(0, 20))
        
        # Test type description
        self.test_type_desc = ttk.Label(test_type_frame, text="", font=('Arial', 9), foreground="blue")
        self.test_type_desc.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(5, 0))
        
        # File controls (initially hidden) with better spacing
        self.file_controls_frame = ttk.Frame(req_input_frame)
        self.file_controls_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(15, 10))
        self.file_controls_frame.columnconfigure(1, weight=1)
        
        # File selection row with proper spacing
        ttk.Label(self.file_controls_frame, text="File:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.file_entry = ttk.Entry(self.file_controls_frame, textvariable=self.requirement_path, state='readonly')
        self.file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 15))
        
        browse_btn = ttk.Button(self.file_controls_frame, text="📁 Browse...", command=self.browse_file)
        browse_btn.grid(row=0, column=2, padx=(0, 0))
        
        # File status label for better feedback with proper spacing
        self.file_status_label = ttk.Label(self.file_controls_frame, text="", foreground="green", font=('Arial', 9))
        self.file_status_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
        
        # Action buttons
        action_buttons_frame = ttk.Frame(req_input_frame)
        action_buttons_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(15, 15))
        
        self.load_example_btn = ttk.Button(action_buttons_frame, text="📋 Load Example", 
                                          command=self.load_example_requirement)
        self.load_example_btn.pack(side=tk.LEFT, padx=(0, 20))
        
        self.clear_btn = ttk.Button(action_buttons_frame, text="🗑️ Clear", 
                                   command=self.clear_requirement)
        self.clear_btn.pack(side=tk.LEFT)
        
        # Requirement text area
        text_frame = ttk.Frame(req_input_frame)
        text_frame.grid(row=4, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(15, 0))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        # Add label for the text area
        ttk.Label(text_frame, text="📝 Requirements:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        self.req_text = scrolledtext.ScrolledText(text_frame, height=8, wrap=tk.WORD,
                                                 font=('Consolas', 11), bg='white', fg='black',
                                                 insertbackground='black')
        self.req_text.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Bind mouse wheel scrolling to text widget
        self.bind_mousewheel_to_widget(self.req_text)
        
        # Add placeholder text
        placeholder_text = """Enter your requirements here...

Example:
User Login System:
- Users can log in with email and password
- Invalid credentials show appropriate error message  
- Password reset functionality available
- Account lockout after multiple failed attempts
- Remember me option for convenience

You can also load requirements from a file and edit them here."""
        
        self.req_text.insert(1.0, placeholder_text)
        self.req_text.config(fg='gray', bg='white')  # Set initial placeholder colors
        self.req_text.bind('<FocusIn>', self.on_text_focus_in)
        self.req_text.bind('<KeyPress>', self.on_text_changed)
        
        # Initially hide file controls
        self.file_controls_frame.grid_remove()
        
        # Generation buttons with improved visual hierarchy
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=(15, 20))
        
        # Secondary action button (Enhance)
        self.enhance_btn = ttk.Button(btn_frame, text="📝 Enhance Requirement", 
                                     command=self.enhance_requirement)
        self.enhance_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Primary action button (Generate) - make it more prominent
        self.generate_btn = ttk.Button(btn_frame, text="🚀 Generate Test Cases", 
                                      command=self.generate_test_cases)
        self.generate_btn.pack(side=tk.LEFT, padx=(0, 15))
        
        # Configure primary button style for prominence
        try:
            style.configure("Primary.TButton", 
                           font=(default_font[0], default_font[1] + 2, "bold"))
            # Try to make it more visually prominent
            style.map("Primary.TButton",
                     background=[('active', '#0078d4'), ('pressed', '#106ebe')])
            self.generate_btn.configure(style="Primary.TButton")
        except Exception:
            # Fallback if custom styling fails
            self.generate_btn.configure(style="Accent.TButton")
        
        # Progress indication - make it more subtle
        self.progress = ttk.Progressbar(btn_frame, mode='indeterminate', length=150)
        self.progress.pack(side=tk.LEFT, padx=(20, 10))
        
        # Progress status label
        self.progress_label = ttk.Label(btn_frame, text="", font=("Arial", 9))
        self.progress_label.pack(side=tk.LEFT, padx=(5, 0))
        
        # Test cases preview
        preview_frame = ttk.LabelFrame(main_frame, text="📝 Generated Test Cases", padding="10")
        preview_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)
        
        # Header frame with expand/collapse button
        header_frame = ttk.Frame(preview_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        header_frame.columnconfigure(0, weight=1)
        
        # Expand button and info
        self.is_expanded = False
        self.expand_btn = ttk.Button(header_frame, text="� Open in New Window", 
                                    command=self.toggle_expand_view)
        self.expand_btn.grid(row=0, column=0, sticky=tk.W)
        
        # Info label
        self.view_info_label = ttk.Label(header_frame, 
                                        text="💡 Double-click any row for detailed view", 
                                        font=('Arial', 9))
        self.view_info_label.grid(row=0, column=1, sticky=tk.E)
        
        # Treeview for test cases
        columns = ('ID', 'Title', 'Steps', 'Expected', 'Priority', 'Quality')
        self.tree = ttk.Treeview(preview_frame, columns=columns, show='headings', height=8)
        
        # Configure column headings and widths
        self.tree.heading('ID', text='Test ID')
        self.tree.heading('Title', text='Title')
        self.tree.heading('Steps', text='Steps')
        self.tree.heading('Expected', text='Expected Result')
        self.tree.heading('Priority', text='Priority')
        self.tree.heading('Quality', text='Quality Score')
        
        # Improved column widths for better readability
        self.tree.column('ID', width=70, minwidth=70)
        self.tree.column('Title', width=200, minwidth=150)
        self.tree.column('Steps', width=320, minwidth=250)
        self.tree.column('Expected', width=280, minwidth=200)
        self.tree.column('Priority', width=90, minwidth=80)
        self.tree.column('Quality', width=110, minwidth=100)
        
        # Scrollbars for treeview
        tree_scroll_y = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_scroll_y.grid(row=1, column=1, sticky=(tk.N, tk.S))
        tree_scroll_x.grid(row=2, column=0, sticky=(tk.W, tk.E))
        
        # Bind mouse wheel scrolling to tree widget
        self.bind_mousewheel_to_widget(self.tree)
        
        # Bind double-click event for detailed view
        self.tree.bind('<Double-1>', self.show_test_case_detail)
        
        # Quality metrics section
        quality_frame = ttk.LabelFrame(main_frame, text="📊 Quality Assessment", padding="10")
        quality_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        quality_frame.columnconfigure(1, weight=1)
        
        # Overall quality score
        ttk.Label(quality_frame, text="Overall Score:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.quality_score_var = tk.StringVar(value="Not assessed")
        self.quality_score_label = ttk.Label(quality_frame, textvariable=self.quality_score_var, 
                                           font=('Arial', 12, 'bold'))
        self.quality_score_label.grid(row=0, column=1, sticky=tk.W, padx=(0, 10))
        
        # Quality insights button
        self.quality_btn = ttk.Button(quality_frame, text="📋 View Quality Report", 
                                     command=self.show_quality_report, state='disabled')
        self.quality_btn.grid(row=0, column=2, padx=(0, 5))
        
        # Quality distribution
        quality_dist_frame = ttk.Frame(quality_frame)
        quality_dist_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        self.quality_dist_var = tk.StringVar(value="")
        self.quality_dist_label = ttk.Label(quality_dist_frame, textvariable=self.quality_dist_var)
        self.quality_dist_label.pack(side=tk.LEFT)
        
        # Actions section
        actions_frame = ttk.LabelFrame(main_frame, text="🔗 Actions", padding="10")
        actions_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # TestRail actions
        testrail_subframe = ttk.Frame(actions_frame)
        testrail_subframe.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(testrail_subframe, text="TestRail Integration:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        
        testrail_buttons = ttk.Frame(testrail_subframe)
        testrail_buttons.pack(fill=tk.X, pady=(5, 0))
        
        self.approve_btn = ttk.Button(testrail_buttons, text="✅ Approve & Push to TestRail", 
                                     command=self.approve_and_push, state='disabled')
        self.approve_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.reject_btn = ttk.Button(testrail_buttons, text="❌ Reject", 
                                    command=self.reject_cases, state='disabled')
        self.reject_btn.pack(side=tk.LEFT)
        
        # Download actions
        download_subframe = ttk.Frame(actions_frame)
        download_subframe.pack(fill=tk.X)
        
        ttk.Label(download_subframe, text="Export Test Cases:", font=('Arial', 9, 'bold')).pack(anchor=tk.W)
        
        download_buttons = ttk.Frame(download_subframe)
        download_buttons.pack(fill=tk.X, pady=(5, 0))
        
        self.download_csv_btn = ttk.Button(download_buttons, text="📄 Download as CSV", 
                                          command=self.download_csv, state='disabled')
        self.download_csv_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.download_xlsx_btn = ttk.Button(download_buttons, text="📊 Download as Excel", 
                                           command=self.download_xlsx, state='disabled')
        self.download_xlsx_btn.pack(side=tk.LEFT)
        
        # Log display
        log_frame = ttk.LabelFrame(main_frame, text="📊 Activity Log", padding="10")
        log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, wrap=tk.WORD)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    def load_default_requirement(self):
        """Load the first available requirement file."""
        try:
            req_files = sorted(self.REQ_DIR.glob("*.txt"))
            if req_files:
                self.load_requirement_file(req_files[0])
                self.logger.info(f"Loaded default requirement: {req_files[0].name}")
        except Exception as e:
            self.logger.warning(f"Could not load default requirement: {e}")
    
    def on_input_method_change(self):
        """Handle input method radio button changes."""
        method = self.input_method.get()
        
        if method == "file":
            # Show file controls
            self.file_controls_frame.grid()
            self.load_example_btn.config(state='disabled')
        else:
            # Hide file controls
            self.file_controls_frame.grid_remove()
            self.load_example_btn.config(state='normal')
            
        self.logger.info(f"Input method changed to: {method}")
    
    def on_test_type_change(self):
        """Handle test type selection changes."""
        test_type = self.test_type.get()
        
        # Update description based on test type
        descriptions = {
            "smoke": "Critical functionality tests - Verify core features work",
            "sanity": "Focused sanity tests - Verify specific functionality after changes", 
            "unit": "Code-level tests - Verify individual functions/methods work correctly"
        }
        
        self.test_type_desc.config(text=descriptions.get(test_type, ""))
        
        # Update placeholder text if it's currently showing
        if self.is_placeholder_text:
            self.update_placeholder_text()
            
        self.logger.info(f"Test type changed to: {test_type}")
    
    def update_placeholder_text(self):
        """Update placeholder text based on selected test type."""
        test_type = self.test_type.get()
        
        placeholders = {
            "smoke": """Enter your requirements here...

Example for Smoke Tests:
User Login System:
- Users can log in with email and password
- Dashboard loads after successful login
- Critical navigation menu is accessible
- User can log out successfully
- Basic user profile information displays correctly

Focus on core, critical functionality that must work for the system to be viable.""",

            "sanity": """Enter your requirements here...

Example for Sanity Tests:
Recent Password Reset Feature Changes:
- Password reset email is sent successfully
- Reset link in email works correctly
- New password meets complexity requirements
- User can login with new password
- Old password is no longer valid
- Password reset link expires after use

Focus on recently changed or specific functionality areas.""",

            "unit": """Enter your code here for unit test generation...

Example Code for Unit Tests:
```python
def calculate_discount(price, discount_percent, user_type):
    \"\"\"Calculate discount amount for a given price.
    
    Args:
        price (float): Original price
        discount_percent (int): Discount percentage (0-100)
        user_type (str): Type of user ('regular', 'premium', 'vip')
    
    Returns:
        float: Discounted price
    \"\"\"
    if price < 0:
        raise ValueError("Price cannot be negative")
    
    if discount_percent < 0 or discount_percent > 100:
        raise ValueError("Discount must be between 0 and 100")
    
    # Apply additional discount for premium users
    if user_type == 'premium':
        discount_percent += 5
    elif user_type == 'vip':
        discount_percent += 10
    
    # Cap discount at 90%
    discount_percent = min(discount_percent, 90)
    
    discount_amount = price * (discount_percent / 100)
    return price - discount_amount
```

Paste your function/class code above for comprehensive unit test generation."""
        }
        
        # Clear and update placeholder text
        self.req_text.delete(1.0, tk.END)
        self.req_text.insert(1.0, placeholders.get(test_type, placeholders["smoke"]))
        self.req_text.config(fg='gray', bg='white')
        self.is_placeholder_text = True
    
    def on_text_focus_in(self, event):
        """Handle text area focus to remove placeholder."""
        if self.is_placeholder_text:
            self.req_text.delete(1.0, tk.END)
            self.req_text.config(fg='black', bg='white')
            self.is_placeholder_text = False
    
    def on_text_changed(self, event):
        """Handle text changes to track if content is placeholder."""
        if self.is_placeholder_text:
            self.is_placeholder_text = False
            self.req_text.config(fg='black', bg='white')
    
    def load_example_requirement(self):
        """Load an example requirement for demonstration."""
        example_requirements = [
            """E-Commerce Shopping Cart System:

**Core Functionality:**
- Users can browse products and add items to cart
- Cart displays item details, quantities, and total price
- Users can update quantities or remove items from cart
- Apply discount codes and promotional offers
- Calculate shipping costs based on location
- Support multiple payment methods (credit card, PayPal, digital wallets)

**User Experience Requirements:**
- Cart contents persist across browser sessions
- Real-time inventory validation before checkout
- Clear error messages for invalid operations
- Mobile-responsive design for all devices
- Loading indicators for async operations

**Security & Validation:**
- Validate all user inputs and sanitize data
- Secure payment processing with encryption
- Session management and timeout handling
- Prevent cart manipulation and price tampering

**Performance Requirements:**
- Cart operations should respond within 2 seconds
- Support concurrent users without conflicts
- Graceful handling of network connectivity issues""",

            """User Authentication & Account Management:

**Login System:**
- Users can log in with email/username and password
- Support for social media login (Google, Facebook, Apple)
- Two-factor authentication for enhanced security
- Remember me functionality with secure tokens
- Account lockout after multiple failed attempts

**Registration Process:**
- User registration with email verification
- Password strength requirements and validation
- Terms and conditions acceptance
- Optional profile information collection
- Welcome email with account setup instructions

**Password Management:**
- Forgot password functionality with email reset
- Secure password reset with time-limited tokens
- Password change with current password verification
- Password history to prevent reuse of recent passwords

**Account Features:**
- User profile management and editing
- Account deactivation and deletion options
- Login history and security activity logs
- Privacy settings and notification preferences""",

            """Flight Booking System:

**Search Functionality:**
- Search flights by origin, destination, and travel dates
- Support for round-trip, one-way, and multi-city trips
- Filter results by price, duration, airline, and stops
- Sort options by price, duration, departure time
- Flexible date search with calendar view

**Booking Process:**
- Select flights and view detailed itinerary
- Passenger information entry with validation
- Seat selection with interactive seat map
- Add-on services (baggage, meals, insurance)
- Payment processing with multiple options
- Booking confirmation with reference number

**Booking Management:**
- View booking details and itinerary
- Modify bookings (date changes, passenger details)
- Cancel bookings with refund processing
- Check-in functionality with mobile boarding pass
- Flight status updates and notifications

**Integration Requirements:**
- Real-time flight data from airline systems
- Payment gateway integration for secure transactions
- Email and SMS notifications for booking updates
- Integration with loyalty programs and frequent flyer accounts"""
        ]
        
        import random
        selected_example = random.choice(example_requirements)
        
        self.req_text.delete(1.0, tk.END)
        self.req_text.insert(1.0, selected_example)
        self.req_text.config(fg='black', bg='white')
        self.is_placeholder_text = False
        
        # Clear any previous file selection
        self.requirement_path.set("")
        
        self.logger.info("📋 Loaded example requirement")
    
    def clear_requirement(self):
        """Clear the requirement text area."""
        self.req_text.delete(1.0, tk.END)
        self.requirement_path.set("")
        self.requirement_text = ""
        self.original_requirement_text = ""
        self.is_placeholder_text = True
        
        # Restore placeholder text
        placeholder_text = """Enter your requirements here...

Example:
User Login System:
- Users can log in with email and password
- Invalid credentials show appropriate error message  
- Password reset functionality available
- Account lockout after multiple failed attempts
- Remember me option for convenience

You can also load requirements from a file and edit them here."""
        
        self.req_text.insert(1.0, placeholder_text)
        self.req_text.config(fg='gray', bg='white')
        
        # Clear file status
        self.file_status_label.config(text="")
        
        # Clear previous results
        self.clear_test_cases()
        self.enhancement_report = {}
        
        self.logger.info("🗑️ Cleared requirement text")
    
    def toggle_expand_view(self):
        """Open expanded view in a new window for better usability."""
        if not self.generated_cases:
            messagebox.showinfo("Info", "No test cases to display. Please generate test cases first.")
            return
        
        # Create expanded view window
        self.open_expanded_view_window()
        self.logger.info("📊 Opened expanded view in new window")
    
    def show_test_case_detail(self, event):
        """Show detailed view of selected test case in a popup window."""
        self.show_test_case_detail_from_tree(event, self.tree)
    
    def open_expanded_view_window(self):
        """Open a new window with expanded test cases view for better usability."""
        # Create expanded view window
        expanded_window = tk.Toplevel(self.root)
        expanded_window.title("📊 Test Cases - Expanded View")
        expanded_window.geometry("1200x700")
        expanded_window.resizable(True, True)
        
        # Make it stay on top initially but allow user to manage it
        expanded_window.transient(self.root)
        
        # Main frame with padding
        main_frame = ttk.Frame(expanded_window, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        expanded_window.columnconfigure(0, weight=1)
        expanded_window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # Header frame with title and controls
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.columnconfigure(1, weight=1)
        
        # Title and info
        title_label = ttk.Label(header_frame, text="📝 All Test Cases - Detailed View", 
                               font=('Arial', 12, 'bold'))
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        info_label = ttk.Label(header_frame, 
                              text="💡 Double-click any row for full detailed view • Resize window as needed", 
                              font=('Arial', 9))
        info_label.grid(row=0, column=1, sticky=tk.E)
        
        # Test cases count
        cases_count = len(self.generated_cases)
        count_label = ttk.Label(header_frame, text=f"Total Test Cases: {cases_count}", 
                               font=('Arial', 10))
        count_label.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        
        # Close button
        close_btn = ttk.Button(header_frame, text="✖️ Close Window", 
                              command=expanded_window.destroy)
        close_btn.grid(row=1, column=1, sticky=tk.E, pady=(5, 0))
        
        # Treeview frame
        tree_frame = ttk.Frame(main_frame)
        tree_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
        # Create expanded treeview with more generous sizing
        columns = ('ID', 'Title', 'Steps', 'Expected', 'Priority', 'Quality')
        expanded_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        
        # Configure column headings
        expanded_tree.heading('ID', text='Test ID')
        expanded_tree.heading('Title', text='Test Case Title')
        expanded_tree.heading('Steps', text='Test Steps')
        expanded_tree.heading('Expected', text='Expected Results')
        expanded_tree.heading('Priority', text='Priority')
        expanded_tree.heading('Quality', text='Quality Score')
        
        # Configure column widths for better readability in expanded view
        expanded_tree.column('ID', width=80, minwidth=80)
        expanded_tree.column('Title', width=250, minwidth=200)
        expanded_tree.column('Steps', width=400, minwidth=300)
        expanded_tree.column('Expected', width=350, minwidth=250)
        expanded_tree.column('Priority', width=100, minwidth=80)
        expanded_tree.column('Quality', width=120, minwidth=100)
        
        # Scrollbars for the expanded treeview
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=expanded_tree.yview)
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=expanded_tree.xview)
        expanded_tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        
        expanded_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_scroll_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        tree_scroll_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Bind double-click for detailed view
        expanded_tree.bind('<Double-1>', lambda event: self.show_test_case_detail_from_tree(event, expanded_tree))
        
        # Populate the expanded treeview
        self._populate_expanded_tree(expanded_tree)
        
        # Focus the window
        expanded_window.focus_set()
        
        self.logger.info(f"📊 Opened expanded view window with {cases_count} test cases")
    
    def show_test_case_detail_from_tree(self, event, tree_widget):
        """Show detailed view of selected test case from any tree widget."""
        selection = tree_widget.selection()
        if not selection:
            return
        
        item = tree_widget.item(selection[0])
        values = item['values']
        
        if not values:
            return
        
        # Extract test case details
        test_id = values[0]
        title = values[1] 
        steps = values[2]
        expected = values[3]
        priority = values[4]
        quality = values[5] if len(values) > 5 else "N/A"
        
        # Find the full test case data for complete details
        full_case = None
        for case in self.generated_cases:
            if case.get("id", "") == test_id:
                full_case = case
                break
        
        if full_case:
            # Use full data for the detailed view
            self._show_detailed_test_case(test_id, full_case.get("title", ""), 
                                        full_case.get("steps", []), full_case.get("expected", ""),
                                        full_case.get("priority", "Medium"), quality)
        else:
            # Fallback to truncated data
            self._show_detailed_test_case(test_id, title, steps, expected, priority, quality)
    
    def _show_detailed_test_case(self, test_id, title, steps, expected, priority, quality):
        """Show detailed test case information in a popup window."""
        # Create detail window
        detail_window = tk.Toplevel(self.root)
        detail_window.title(f"Test Case Details - {test_id}")
        detail_window.geometry("800x600")
        detail_window.resizable(True, True)
        
        # Make it modal
        detail_window.transient(self.root)
        detail_window.grab_set()
        
        # Main frame with padding
        main_frame = ttk.Frame(detail_window, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        detail_window.columnconfigure(0, weight=1)
        detail_window.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Test ID and Quality
        ttk.Label(main_frame, text="Test ID:", font=('Arial', 10, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
        ttk.Label(main_frame, text=str(test_id), font=('Arial', 10)).grid(row=0, column=1, sticky=tk.W, pady=(0, 10))
        
        ttk.Label(main_frame, text="Quality Score:", font=('Arial', 10, 'bold')).grid(row=0, column=2, sticky=tk.W, padx=(20, 0), pady=(0, 10))
        ttk.Label(main_frame, text=str(quality), font=('Arial', 10)).grid(row=0, column=3, sticky=tk.W, pady=(0, 10))
        
        # Title
        ttk.Label(main_frame, text="Title:", font=('Arial', 10, 'bold')).grid(row=1, column=0, sticky=(tk.W, tk.N), pady=(0, 10))
        title_text = tk.Text(main_frame, height=2, wrap=tk.WORD, font=('Arial', 10), bg='#f0f0f0', fg='black')
        title_text.grid(row=1, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        title_text.insert(1.0, str(title))
        title_text.config(state='disabled')
        
        # Priority
        ttk.Label(main_frame, text="Priority:", font=('Arial', 10, 'bold')).grid(row=2, column=0, sticky=tk.W, pady=(0, 10))
        ttk.Label(main_frame, text=str(priority), font=('Arial', 10)).grid(row=2, column=1, sticky=tk.W, pady=(0, 10))
        
        # Test Steps
        ttk.Label(main_frame, text="Test Steps:", font=('Arial', 10, 'bold')).grid(row=3, column=0, sticky=(tk.W, tk.N), pady=(0, 5))
        
        steps_frame = ttk.Frame(main_frame)
        steps_frame.grid(row=4, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        steps_frame.columnconfigure(0, weight=1)
        steps_frame.rowconfigure(0, weight=1)
        
        steps_text = tk.Text(steps_frame, height=8, wrap=tk.WORD, font=('Arial', 10), bg='white', fg='black')
        steps_scroll = ttk.Scrollbar(steps_frame, orient=tk.VERTICAL, command=steps_text.yview)
        steps_text.configure(yscrollcommand=steps_scroll.set)
        
        steps_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        steps_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Format steps properly
        if isinstance(steps, list):
            formatted_steps = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps)])
        else:
            formatted_steps = str(steps)
        steps_text.insert(1.0, formatted_steps)
        steps_text.config(state='disabled')
        
        # Expected Results
        ttk.Label(main_frame, text="Expected Results:", font=('Arial', 10, 'bold')).grid(row=5, column=0, sticky=(tk.W, tk.N), pady=(0, 5))
        
        expected_frame = ttk.Frame(main_frame)
        expected_frame.grid(row=6, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        expected_frame.columnconfigure(0, weight=1)
        expected_frame.rowconfigure(0, weight=1)
        
        expected_text = tk.Text(expected_frame, height=6, wrap=tk.WORD, font=('Arial', 10), bg='white', fg='black')
        expected_scroll = ttk.Scrollbar(expected_frame, orient=tk.VERTICAL, command=expected_text.yview)
        expected_text.configure(yscrollcommand=expected_scroll.set)
        
        expected_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        expected_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        expected_text.insert(1.0, str(expected))
        expected_text.config(state='disabled')
        
        # Configure row weights for proper expansion
        main_frame.rowconfigure(4, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        # Close button
        close_btn = ttk.Button(main_frame, text="✖️ Close", 
                              command=detail_window.destroy)
        close_btn.grid(row=7, column=0, columnspan=4, pady=(10, 0))
        
        # Focus the window
        detail_window.focus_set()
        
        self.logger.info(f"📖 Opened detailed view for test case: {test_id}")
    
    def _populate_expanded_tree(self, tree_widget):
        """Populate the expanded tree with test case data."""
        # Create quality score mapping for easy lookup
        quality_scores = {}
        if self.quality_report and "individual_scores" in self.quality_report:
            for score_info in self.quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                quality_scores[test_id] = total_score
        
        # Populate treeview with more generous text limits for expanded view
        for case in self.generated_cases:
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = " | ".join(steps)
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            # For expanded view, show more text but still reasonable limits
            title_display = case.get("title", "")[:60] + "..." if len(case.get("title", "")) > 60 else case.get("title", "")
            steps_display = steps_text[:300] + "..." if len(steps_text) > 300 else steps_text
            expected_display = case.get("expected", "")[:200] + "..." if len(case.get("expected", "")) > 200 else case.get("expected", "")
            
            item = tree_widget.insert('', tk.END, values=(
                test_id,
                title_display,
                steps_display,
                expected_display,
                case.get("priority", "Medium"),
                quality_display
            ))
            
            # Apply color coding to quality column if needed
            if quality_score > 0:
                tree_widget.set(item, 'Quality', quality_display)
    
    def _refresh_tree_display(self):
        """Refresh the tree display with current cases and expanded state."""
        if not self.generated_cases:
            return
            
        # Clear and repopulate tree
        self.clear_test_cases()
        
        # Create quality score mapping for easy lookup
        quality_scores = {}
        if self.quality_report and "individual_scores" in self.quality_report:
            for score_info in self.quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                quality_scores[test_id] = total_score
        
        # Populate treeview with appropriate text length
        for case in self.generated_cases:
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = " | ".join(steps)
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            # Handle text display based on expanded state
            if self.is_expanded:
                # In expanded view, show more text but still limit for readability
                steps_display = steps_text[:200] + "..." if len(steps_text) > 200 else steps_text
                expected_display = case.get("expected", "")[:150] + "..." if len(case.get("expected", "")) > 150 else case.get("expected", "")
                title_display = case.get("title", "")[:50] + "..." if len(case.get("title", "")) > 50 else case.get("title", "")
            else:
                # In normal view, show truncated text
                steps_display = steps_text[:80] + "..." if len(steps_text) > 80 else steps_text
                expected_display = case.get("expected", "")[:100] + "..." if len(case.get("expected", "")) > 100 else case.get("expected", "")
                title_display = case.get("title", "")[:40] + "..." if len(case.get("title", "")) > 40 else case.get("title", "")
            
            item = self.tree.insert('', tk.END, values=(
                test_id,
                title_display,
                steps_display,
                expected_display,
                case.get("priority", "Medium"),
                quality_display
            ))
            
            # Apply color coding to quality column
            if quality_score > 0:
                self.tree.set(item, 'Quality', quality_display)
    
    def browse_file(self):
        """Open file dialog to select requirement file."""
        filename = filedialog.askopenfilename(
            title="Select Requirement File",
            initialdir=self.REQ_DIR,
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if filename:
            self.load_requirement_file(Path(filename))
    
    def load_requirement_file(self, file_path):
        """Load and display requirement file content."""
        try:
            self.requirement_path.set(str(file_path))
            self.requirement_text = file_path.read_text(encoding="utf-8").strip()
            self.original_requirement_text = self.requirement_text  # Store original
            
            # Switch to file mode
            self.input_method.set("file")
            self.on_input_method_change()
            
            # Display in text widget
            self.req_text.delete(1.0, tk.END)
            self.req_text.insert(1.0, self.requirement_text)
            self.req_text.config(fg='black', bg='white')
            self.is_placeholder_text = False
            
            # Show file loaded confirmation
            self.file_status_label.config(text=f"✅ Loaded: {file_path.name}")
            
            # Clear previous results and reports
            self.clear_test_cases()
            self.enhancement_report = {}
            
            self.logger.info(f"📁 Loaded requirement from file: {file_path.name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load requirement file:\n{e}")
    
    def generate_test_cases(self):
        """Generate test cases using LLM in a separate thread."""
        if self.is_processing:
            return
        
        # Get current requirement text from the text area
        current_text = self.req_text.get(1.0, tk.END).strip()
        
        # Check if we have actual content (not placeholder)
        if not current_text or self.is_placeholder_text or current_text.startswith("Enter your requirements here"):
            messagebox.showwarning("Warning", "Please enter requirements or load from a file first.")
            return
        
        # Update the requirement text with current content
        self.requirement_text = current_text
        
        # Start generation in background thread
        self.is_processing = True
        self.generate_btn.config(state='disabled')
        self.progress.start(10)
        self.progress_label.config(text="Generating test cases...")
        
        thread = threading.Thread(target=self._generate_test_cases_thread)
        thread.daemon = True
        thread.start()
    
    def _generate_test_cases_thread(self):
        """Background thread for test case generation."""
        try:
            self.logger.info("🤖 Starting test case generation...")
            
            # Update progress
            self.root.after(0, lambda: self.progress_label.config(text="Analyzing requirements..."))
            
            # Get test type for customized prompts
            test_type = self.test_type.get()
            
            # Customize system prompt based on test type
            test_type_instructions = {
                "smoke": """
Focus on SMOKE TESTS - Critical functionality that must work for basic system operation:
- Core login/authentication flows
- Essential business operations
- Critical user journeys
- System startup and basic navigation
- Key integrations that if broken, make system unusable
Generate 8-12 test cases covering the most critical paths.""",

                "sanity": """
Focus on SANITY TESTS - Targeted testing after changes or in specific areas:
- Recently modified functionality
- Specific feature areas mentioned in requirements
- Quick verification of key functionality
- Regression testing for critical areas
- Narrow but deep testing of mentioned features
Generate 6-10 focused test cases for the specific areas mentioned.""",

                "unit": """
Focus on UNIT TESTS - Code-level testing of individual functions/methods:
- Test each function with valid inputs
- Test boundary conditions and edge cases
- Test error handling and exception cases
- Test different parameter combinations
- Verify return values and side effects
- Mock external dependencies if needed
Generate 10-15 comprehensive unit tests covering all code paths."""
            }
            
            # Enhanced system prompt with test type context
            enhanced_system_prompt = f"""{self.SYSTEM_PROMPT}

{test_type_instructions.get(test_type, test_type_instructions['smoke'])}

Test Type: {test_type.upper()} TESTS"""
            
            # Prepare messages
            user_prompt = self.USER_PROMPT_TEMPLATE.format(requirement_text=self.requirement_text)
            messages = [
                {"role": "system", "content": enhanced_system_prompt},
                {"role": "user", "content": user_prompt}
            ]
            
            # Call LLM
            self.logger.info("📡 Calling LLM API...")
            self.root.after(0, lambda: self.progress_label.config(text="Generating test cases..."))
            raw = chat(messages)
            
            # Parse response
            self.logger.info("📝 Parsing LLM response...")
            self.root.after(0, lambda: self.progress_label.config(text="Processing results..."))
            cases = parse_json_safely(raw, self.LAST_RAW_JSON)
            
            if not cases:
                self.logger.warning("⚠️ No valid test cases generated, using fallback cases")
                cases = [
                    {"id": "TC-001", "title": "Login with valid credentials", 
                     "steps": ["Enter username", "Enter password", "Click login"], 
                     "expected": "User is logged in", "priority": "High"},
                    {"id": "TC-002", "title": "Login with invalid password", 
                     "steps": ["Enter username", "Enter wrong password", "Click login"], 
                     "expected": "Error message displayed", "priority": "High"}
                ]
            
            # Save to CSV
            rows = to_rows(cases)
            write_csv(rows, self.OUT_CSV)
            
            # Perform quality assessment
            self.logger.info("📊 Assessing test case quality...")
            try:
                quality_report = score_test_cases(cases, self.requirement_text, self.OUT_DIR)
                self.logger.info(f"🔍 Quality Assessment Debug - Report generated: {bool(quality_report)}")
                if quality_report:
                    self.logger.info(f"🔍 Quality Assessment Debug - Report keys: {list(quality_report.keys())}")
            except Exception as e:
                self.logger.error(f"❌ Quality assessment failed: {e}")
                # Create a basic quality report as fallback
                quality_report = self._create_fallback_quality_report(cases)
            
            # Update UI in main thread
            self.root.after(0, self._update_ui_with_cases, cases, quality_report)
            
        except Exception as e:
            self.logger.error(f"❌ Generation failed: {e}")
            self.root.after(0, self._generation_error, str(e))
    
    def _update_ui_with_cases(self, cases, quality_report=None):
        """Update UI with generated test cases and quality assessment (called from main thread)."""
        self.generated_cases = cases
        self.quality_report = quality_report or {}
        
        # Debug logging
        self.logger.info(f"🔍 Update UI Debug - Quality report received: {bool(quality_report)}")
        self.logger.info(f"🔍 Update UI Debug - Quality report set: {bool(self.quality_report)}")
        
        # Clear previous results (just the tree, not quality report)
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Create quality score mapping for easy lookup
        quality_scores = {}
        if quality_report and "individual_scores" in quality_report:
            for score_info in quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                quality_scores[test_id] = total_score
        
        # Populate treeview with quality scores
        for case in cases:
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = " | ".join(steps)
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            # Color code quality score
            quality_color = self._get_quality_color(quality_score)
            
            # Handle text display based on expanded state
            if self.is_expanded:
                # In expanded view, show more text but still limit for readability
                steps_display = steps_text[:200] + "..." if len(steps_text) > 200 else steps_text
                expected_display = case.get("expected", "")[:150] + "..." if len(case.get("expected", "")) > 150 else case.get("expected", "")
                title_display = case.get("title", "")[:50] + "..." if len(case.get("title", "")) > 50 else case.get("title", "")
            else:
                # In normal view, show truncated text
                steps_display = steps_text[:80] + "..." if len(steps_text) > 80 else steps_text
                expected_display = case.get("expected", "")[:100] + "..." if len(case.get("expected", "")) > 100 else case.get("expected", "")
                title_display = case.get("title", "")[:40] + "..." if len(case.get("title", "")) > 40 else case.get("title", "")
            
            item = self.tree.insert('', tk.END, values=(
                test_id,
                title_display,
                steps_display,
                expected_display,
                case.get("priority", "Medium"),
                quality_display
            ))
            
            # Apply color coding to quality column
            if quality_score > 0:
                self.tree.set(item, 'Quality', quality_display)
        
        # Update quality metrics display
        self._update_quality_display()
        
        # Enable approval and download buttons
        self.approve_btn.config(state='normal')
        self.reject_btn.config(state='normal')
        self.download_csv_btn.config(state='normal')
        self.download_xlsx_btn.config(state='normal' if EXCEL_AVAILABLE else 'disabled')
        
        # Ensure quality button is enabled if we have any quality report
        if self.quality_report:
            self.quality_btn.config(state='normal')
            self.logger.info("🔍 Quality button explicitly enabled")
        
        # Log quality summary if available
        if quality_report:
            overall_score = quality_report.get("overall_score", 0)
            self.logger.info(f"✅ Generated {len(cases)} test cases with overall quality score: {overall_score:.1f}/10")
        else:
            self.logger.info(f"✅ Generated {len(cases)} test cases successfully!")
        
        self._stop_processing()
    
    def enhance_requirement(self):
        """Enhance the current requirement text for better test case generation."""
        if self.is_processing:
            return
        
        # Get current text from the text widget (in case user edited it)
        current_text = self.req_text.get(1.0, tk.END).strip()
        
        if not current_text:
            messagebox.showwarning("Warning", "Please load a requirement file first.")
            return
        
        # Start enhancement in background thread
        self.is_processing = True
        self.enhance_btn.config(state='disabled')
        self.generate_btn.config(state='disabled')
        self.progress.start(10)
        
        thread = threading.Thread(target=self._enhance_requirement_thread, args=(current_text,))
        thread.daemon = True
        thread.start()
    
    def _enhance_requirement_thread(self, requirement_text):
        """Background thread for requirement enhancement."""
        try:
            self.logger.info("📝 Enhancing requirement for better test case generation...")
            
            # Enhance the requirement
            enhanced_text, report = enhance_requirement(requirement_text, self.OUT_DIR)
            
            # Update UI in main thread
            self.root.after(0, self._update_ui_with_enhancement, enhanced_text, report)
            
        except Exception as e:
            self.logger.error(f"❌ Requirement enhancement failed: {e}")
            self.root.after(0, self._enhancement_error, str(e))
    
    def _update_ui_with_enhancement(self, enhanced_text, report):
        """Update UI with enhanced requirement (called from main thread)."""
        self.requirement_text = enhanced_text
        self.enhancement_report = report
        
        # Update text widget with enhanced requirement
        self.req_text.delete(1.0, tk.END)
        self.req_text.insert(1.0, enhanced_text)
        self.req_text.config(fg='black', bg='white')
        self.is_placeholder_text = False
        
        # Log enhancement summary
        overall_score = report.get("overall_score", 0)
        improvements_count = len(report.get("improvements_made", []))
        
        self.logger.info(f"✅ Requirement enhanced! Score: {overall_score:.1f}/10, {improvements_count} improvements made")
        
        # Show enhancement summary dialog
        self._show_enhancement_summary(report)
        
        self._stop_enhancement_processing()
    
    def _enhancement_error(self, error_msg):
        """Handle enhancement error (called from main thread)."""
        messagebox.showerror("Enhancement Error", f"Failed to enhance requirement:\n{error_msg}")
        self._stop_enhancement_processing()
    
    def _stop_enhancement_processing(self):
        """Stop enhancement processing indicators."""
        self.is_processing = False
        self.enhance_btn.config(state='normal')
        self.generate_btn.config(state='normal')
        self.progress.stop()
    
    def _show_enhancement_summary(self, report):
        """Show enhancement summary in a dialog."""
        if not report:
            return
        
        # Create summary window
        summary_window = tk.Toplevel(self.root)
        summary_window.title("📝 Requirement Enhancement Summary")
        summary_window.geometry("700x500")
        summary_window.transient(self.root)
        summary_window.grab_set()
        
        # Summary content
        summary_frame = ttk.Frame(summary_window, padding="10")
        summary_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(summary_frame, text="📊 Enhancement Summary", 
                               font=('Arial', 14, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Scores frame
        scores_frame = ttk.LabelFrame(summary_frame, text="Quality Scores", padding="10")
        scores_frame.pack(fill=tk.X, pady=(0, 10))
        
        overall_score = report.get("overall_score", 0)
        clarity_score = report.get("clarity_score", 0)
        completeness_score = report.get("completeness_score", 0)
        testability_score = report.get("testability_score", 0)
        
        scores_text = f"""
Overall Score: {overall_score:.1f}/10
Clarity: {clarity_score:.1f}/10  |  Completeness: {completeness_score:.1f}/10  |  Testability: {testability_score:.1f}/10
        """.strip()
        
        scores_label = ttk.Label(scores_frame, text=scores_text, font=('Courier', 10))
        scores_label.pack()
        
        # Improvements section
        improvements = report.get("improvements_made", [])
        if improvements:
            imp_frame = ttk.LabelFrame(summary_frame, text="Improvements Made", padding="10")
            imp_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            
            imp_text = scrolledtext.ScrolledText(imp_frame, height=6, wrap=tk.WORD)
            imp_text.pack(fill=tk.BOTH, expand=True)
            
            for i, improvement in enumerate(improvements, 1):
                imp_text.insert(tk.END, f"{i}. {improvement}\n")
            imp_text.config(state='disabled')
        
        # Recommendations section
        recommendations = report.get("recommended_additions", [])
        if recommendations:
            rec_frame = ttk.LabelFrame(summary_frame, text="Additional Recommendations", padding="10")
            rec_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            
            rec_text = scrolledtext.ScrolledText(rec_frame, height=6, wrap=tk.WORD)
            rec_text.pack(fill=tk.BOTH, expand=True)
            
            for i, rec in enumerate(recommendations, 1):
                rec_text.insert(tk.END, f"{i}. {rec}\n")
            rec_text.config(state='disabled')
        
        # Buttons
        btn_frame = ttk.Frame(summary_window)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="View Original", 
                  command=lambda: self._show_original_requirement()).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(btn_frame, text="Close", 
                  command=summary_window.destroy).pack(side=tk.LEFT)
    
    def _show_original_requirement(self):
        """Show the original requirement text in a dialog."""
        if not self.original_requirement_text:
            messagebox.showinfo("Original Requirement", "No original requirement available.")
            return
        
        # Create comparison window
        comparison_window = tk.Toplevel(self.root)
        comparison_window.title("📄 Original vs Enhanced Requirement")
        comparison_window.geometry("900x600")
        comparison_window.transient(self.root)
        
        # Create notebook for tabs
        notebook = ttk.Notebook(comparison_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Original tab
        original_frame = ttk.Frame(notebook)
        notebook.add(original_frame, text="🔴 Original")
        
        original_text = scrolledtext.ScrolledText(original_frame, wrap=tk.WORD)
        original_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        original_text.insert(1.0, self.original_requirement_text)
        original_text.config(state='disabled')
        
        # Enhanced tab
        enhanced_frame = ttk.Frame(notebook)
        notebook.add(enhanced_frame, text="🟢 Enhanced")
        
        enhanced_text = scrolledtext.ScrolledText(enhanced_frame, wrap=tk.WORD)
        enhanced_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        enhanced_text.insert(1.0, self.requirement_text)
        enhanced_text.config(state='disabled')
        
        # Close button
        ttk.Button(comparison_window, text="Close", 
                  command=comparison_window.destroy).pack(pady=5)
    
    def _get_quality_color(self, score):
        """Get color coding for quality score."""
        if score >= 8.0:
            return "green"
        elif score >= 6.0:
            return "orange" 
        elif score > 0:
            return "red"
        else:
            return "gray"
    
    def _update_quality_display(self):
        """Update the quality metrics display - only enable button after generation."""
        # Debug logging
        self.logger.info(f"🔍 Quality Display Debug - Quality report exists: {bool(self.quality_report)}")
        if self.quality_report:
            self.logger.info(f"🔍 Quality Display Debug - Report keys: {list(self.quality_report.keys())}")
            self.logger.info(f"🔍 Quality Display Debug - Overall score: {self.quality_report.get('overall_score', 'Not found')}")
        
        if not self.quality_report:
            self.quality_score_var.set("Click 'View Quality Report' after generating test cases")
            self.quality_dist_var.set("")
            self.quality_btn.config(state='disabled')
            return
        
        # Show brief indication that quality assessment is ready
        self.quality_score_var.set("Quality assessment completed - click to view details")
        self.quality_dist_var.set("Click 'View Quality Report' button to see detailed metrics")
        
        # Enable quality report button - this is the key change
        self.quality_btn.config(state='normal')
    
    def show_quality_report(self):
        """Show detailed quality report in a popup window."""
        if not self.quality_report:
            messagebox.showinfo("Quality Report", 
                               "No quality assessment available.\n\n" +
                               "Quality assessment may have failed during test case generation. " +
                               "Please check the application logs for more details.")
            return
        
        # Create quality report window
        quality_window = tk.Toplevel(self.root)
        quality_window.title("📊 Test Case Quality Report")
        quality_window.geometry("900x700")
        quality_window.transient(self.root)
        quality_window.grab_set()
        
        # Add overall score header
        header_frame = ttk.Frame(quality_window)
        header_frame.pack(fill=tk.X, padx=10, pady=10)
        
        overall_score = self.quality_report.get("overall_score", 0)
        score_color = "green" if overall_score >= 8.0 else "orange" if overall_score >= 6.0 else "red"
        
        ttk.Label(header_frame, text=f"Overall Quality Score: {overall_score:.1f}/10", 
                 font=('Arial', 14, 'bold'), foreground=score_color).pack()
        
        # Create notebook for different tabs
        notebook = ttk.Notebook(quality_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Summary tab
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="📊 Summary")
        
        summary_text = scrolledtext.ScrolledText(summary_frame, wrap=tk.WORD, height=20)
        summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Generate summary content
        summary_content = self._generate_quality_summary()
        summary_text.insert(1.0, summary_content)
        summary_text.config(state='disabled')
        
        # Individual scores tab
        details_frame = ttk.Frame(notebook)
        notebook.add(details_frame, text="📝 Detailed Scores")
        
        # Create treeview for individual scores
        detail_columns = ('Test ID', 'Clarity', 'Completeness', 'Specificity', 'Testability', 'Coverage', 'Total')
        detail_tree = ttk.Treeview(details_frame, columns=detail_columns, show='headings', height=15)
        
        for col in detail_columns:
            detail_tree.heading(col, text=col)
            detail_tree.column(col, width=100, minwidth=80)
        
        # Populate individual scores
        individual_scores = self.quality_report.get("individual_scores", [])
        for score_info in individual_scores:
            test_id = score_info.get("test_id", "")
            scores = score_info.get("scores", {})
            total_score = score_info.get("total_score", 0)
            
            detail_tree.insert('', tk.END, values=(
                test_id,
                f"{scores.get('clarity', 0):.1f}",
                f"{scores.get('completeness', 0):.1f}",
                f"{scores.get('specificity', 0):.1f}",
                f"{scores.get('testability', 0):.1f}",
                f"{scores.get('coverage', 0):.1f}",
                f"{total_score:.1f}"
            ))
        
        # Add scrollbar to detail tree
        detail_scroll = ttk.Scrollbar(details_frame, orient=tk.VERTICAL, command=detail_tree.yview)
        detail_tree.configure(yscrollcommand=detail_scroll.set)
        
        detail_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0), pady=5)
        detail_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        
        # Recommendations tab
        recommendations_frame = ttk.Frame(notebook)
        notebook.add(recommendations_frame, text="💡 Recommendations")
        
        rec_text = scrolledtext.ScrolledText(recommendations_frame, wrap=tk.WORD, height=20)
        rec_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Generate recommendations content
        rec_content = self._generate_recommendations_content()
        rec_text.insert(1.0, rec_content)
        rec_text.config(state='disabled')
        
        # Close button
        close_btn = ttk.Button(quality_window, text="Close", command=quality_window.destroy)
        close_btn.pack(pady=5)
    
    def _generate_quality_summary(self):
        """Generate quality summary text."""
        if not self.quality_report:
            return "No quality assessment available."
        
        lines = []
        overall_score = self.quality_report.get("overall_score", 0)
        individual_scores = self.quality_report.get("individual_scores", [])
        insights = self.quality_report.get("quality_insights", {})
        
        lines.append("🎯 TEST CASE QUALITY ASSESSMENT REPORT")
        lines.append("=" * 50)
        lines.append("")
        
        lines.append(f"📊 Overall Quality Score: {overall_score:.1f}/10")
        lines.append(f"📝 Total Test Cases Evaluated: {len(individual_scores)}")
        lines.append("")
        
        # Quality distribution
        if individual_scores:
            high_quality = sum(1 for s in individual_scores if s.get("total_score", 0) >= 8.0)
            medium_quality = sum(1 for s in individual_scores if 6.0 <= s.get("total_score", 0) < 8.0)
            low_quality = sum(1 for s in individual_scores if s.get("total_score", 0) < 6.0)
            
            lines.append("🎯 Quality Distribution:")
            lines.append(f"  🟢 High Quality (8.0+): {high_quality} tests")
            lines.append(f"  🟡 Medium Quality (6.0-7.9): {medium_quality} tests")
            lines.append(f"  🔴 Low Quality (<6.0): {low_quality} tests")
            lines.append("")
        
        # Overall feedback
        overall_feedback = insights.get("overall_feedback", "")
        if overall_feedback:
            lines.append("📝 Overall Assessment:")
            lines.append(overall_feedback)
            lines.append("")
        
        # Strengths
        strengths = insights.get("strengths", [])
        if strengths:
            lines.append("✅ Key Strengths:")
            for strength in strengths:
                lines.append(f"  • {strength}")
            lines.append("")
        
        # Coverage gaps
        coverage_gaps = insights.get("coverage_gaps", [])
        if coverage_gaps:
            lines.append("⚠️ Coverage Gaps:")
            for gap in coverage_gaps:
                lines.append(f"  • {gap}")
            lines.append("")
        
        return "\n".join(lines)
    
    def _create_fallback_quality_report(self, cases):
        """Create a basic quality report when AI assessment fails."""
        individual_scores = []
        total_score = 0
        
        for case in cases:
            # Basic scoring based on case completeness
            case_id = case.get('id', 'TC-000')
            title = case.get('title', '')
            steps = case.get('steps', [])
            expected = case.get('expected', '')
            
            # Simple scoring heuristics
            clarity_score = 7.0 if title and len(title) > 10 else 5.0
            completeness_score = 8.0 if steps and len(steps) >= 2 else 6.0
            specificity_score = 7.5 if expected and len(expected) > 20 else 5.5
            testability_score = 7.0  # Default assumption
            coverage_score = 6.5  # Conservative estimate
            
            case_total = (clarity_score + completeness_score + specificity_score + 
                         testability_score + coverage_score) / 5
            total_score += case_total
            
            individual_scores.append({
                "test_id": case_id,
                "total_score": case_total,
                "scores": {
                    "clarity": clarity_score,
                    "completeness": completeness_score,
                    "specificity": specificity_score,
                    "testability": testability_score,
                    "coverage": coverage_score
                },
                "feedback": "Basic assessment (AI scoring unavailable)"
            })
        
        overall_score = total_score / len(cases) if cases else 0
        
        return {
            "overall_score": overall_score,
            "individual_scores": individual_scores,
            "quality_insights": {
                "strengths": ["Test cases generated successfully"],
                "weaknesses": ["Detailed AI assessment unavailable"],
                "suggestions": ["Manual review recommended"]
            },
            "assessment_note": "Fallback assessment - AI quality scoring was unavailable"
        }
    
    def _generate_recommendations_content(self):
        """Generate recommendations content."""
        if not self.quality_report:
            return "No recommendations available."
        
        lines = []
        insights = self.quality_report.get("quality_insights", {})
        individual_scores = self.quality_report.get("individual_scores", [])
        
        lines.append("💡 IMPROVEMENT RECOMMENDATIONS")
        lines.append("=" * 40)
        lines.append("")
        
        # General recommendations
        recommendations = insights.get("recommendations", [])
        if recommendations:
            lines.append("🎯 General Recommendations:")
            for i, rec in enumerate(recommendations, 1):
                lines.append(f"  {i}. {rec}")
            lines.append("")
        
        # Missing categories
        missing_categories = insights.get("missing_categories", [])
        if missing_categories:
            lines.append("📋 Missing Test Categories:")
            for category in missing_categories:
                lines.append(f"  • {category}")
            lines.append("")
        
        # Individual test improvements
        lines.append("🔧 Individual Test Case Improvements:")
        lines.append("")
        
        for score_info in individual_scores:
            test_id = score_info.get("test_id", "")
            suggestions = score_info.get("suggestions", [])
            weaknesses = score_info.get("weaknesses", [])
            total_score = score_info.get("total_score", 0)
            
            if suggestions or weaknesses:
                lines.append(f"📝 {test_id} (Score: {total_score:.1f}/10):")
                
                if weaknesses:
                    lines.append("  Weaknesses:")
                    for weakness in weaknesses:
                        lines.append(f"    - {weakness}")
                
                if suggestions:
                    lines.append("  Suggestions:")
                    for suggestion in suggestions:
                        lines.append(f"    + {suggestion}")
                
                lines.append("")
        
        return "\n".join(lines)
    
    def _generation_error(self, error_msg):
        """Handle generation error (called from main thread)."""
        messagebox.showerror("Generation Error", f"Failed to generate test cases:\n{error_msg}")
        self._stop_processing()
    
    def _stop_processing(self):
        """Stop processing indicators."""
        self.is_processing = False
        self.generate_btn.config(state='normal')
        self.progress.stop()
        self.progress_label.config(text="")
    
    def clear_test_cases(self):
        """Clear test cases display and quality metrics."""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.approve_btn.config(state='disabled')
        self.reject_btn.config(state='disabled')
        self.download_csv_btn.config(state='disabled')
        self.download_xlsx_btn.config(state='disabled')
        
        # Clear quality metrics
        self.quality_score_var.set("Click 'View Quality Report' after generating test cases")
        self.quality_dist_var.set("")
        self.quality_btn.config(state='disabled')
        self.quality_report = {}
    
    def approve_and_push(self):
        """Approve test cases and push to TestRail."""
        if not self.generated_cases:
            return
        
        # Enhanced confirmation dialog with more details
        total_cases = len(self.generated_cases)
        avg_quality = 0
        if self.quality_report and "overall_score" in self.quality_report:
            avg_quality = self.quality_report["overall_score"]
        
        message = f"Push {total_cases} test cases to TestRail?"
        detail = (f"This will create {total_cases} new test cases in your TestRail project.\n"
                 f"Average Quality Score: {avg_quality:.1f}/10\n\n"
                 f"This action cannot be undone. Are you sure you want to continue?")
        
        result = messagebox.askyesno(
            "⚠️  Confirm TestRail Push", 
            message,
            detail=detail
        )
        
        if result:
            # Start push in background thread
            self.approve_btn.config(state='disabled')
            self.progress.start(10)
            
            thread = threading.Thread(target=self._push_to_testrail_thread)
            thread.daemon = True
            thread.start()
    
    def _push_to_testrail_thread(self):
        """Background thread for TestRail push."""
        try:
            self.logger.info("📤 Pushing test cases to TestRail...")
            
            created_ids = []
            for case in self.generated_cases:
                try:
                    payload = map_case_to_testrail_payload(case)
                    res = create_case(payload)
                    cid = res.get("id")
                    if cid:
                        created_ids.append(cid)
                        add_result(cid, status_id=3, comment="Created by Desktop UI")
                        self.logger.info(f"✅ Created TestRail case: {case.get('title', '')}")
                except Exception as e:
                    self.logger.error(f"❌ Failed to create case '{case.get('title', '')}': {e}")
            
            # Update UI
            self.root.after(0, self._push_complete, created_ids)
            
        except Exception as e:
            self.logger.error(f"❌ TestRail push failed: {e}")
            self.root.after(0, self._push_error, str(e))
    
    def _push_complete(self, created_ids):
        """Handle successful TestRail push (called from main thread)."""
        self.progress.stop()
        self.logger.info(f"🎉 Successfully pushed {len(created_ids)} test cases to TestRail!")
        
        # Show project stats
        try:
            stats = get_stats()
            total = stats.get("total_cases", 0)
            self.logger.info(f"📊 TestRail project now has {total} total test cases")
        except Exception as e:
            self.logger.warning(f"Could not fetch project stats: {e}")
        
        messagebox.showinfo(
            "Success", 
            f"Successfully created {len(created_ids)} test cases in TestRail!\n\nCase IDs: {created_ids}"
        )
    
    def _push_error(self, error_msg):
        """Handle TestRail push error (called from main thread)."""
        self.progress.stop()
        self.approve_btn.config(state='normal')
        messagebox.showerror("TestRail Error", f"Failed to push to TestRail:\n{error_msg}")
    
    def reject_cases(self):
        """Reject generated test cases."""
        result = messagebox.askyesno(
            "Confirm Rejection", 
            "Reject the generated test cases?",
            detail="This will clear the current test cases. You can generate new ones."
        )
        
        if result:
            self.clear_test_cases()
            self.generated_cases = []
            self.logger.info("🚫 Test cases rejected by user")
    
    def download_csv(self):
        """Download test cases as CSV file."""
        if not self.generated_cases:
            messagebox.showwarning("No Data", "No test cases available to download.")
            return
        
        # Ask user for save location
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"test_cases_{timestamp}.csv"
        
        file_path = filedialog.asksaveasfilename(
            title="Save Test Cases as CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not file_path:
            return
        
        try:
            self._export_to_csv(file_path)
            self.logger.info(f"📄 Test cases exported to CSV: {file_path}")
            messagebox.showinfo("Export Success", f"Test cases successfully exported to:\n{file_path}")
        except Exception as e:
            self.logger.error(f"❌ CSV export failed: {e}")
            messagebox.showerror("Export Error", f"Failed to export CSV:\n{e}")
    
    def download_xlsx(self):
        """Download test cases as Excel file."""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Excel Not Available", 
                               "Excel export requires openpyxl library.\n"
                               "Install it with: pip install openpyxl")
            return
        
        if not self.generated_cases:
            messagebox.showwarning("No Data", "No test cases available to download.")
            return
        
        # Ask user for save location
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"test_cases_{timestamp}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            title="Save Test Cases as Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not file_path:
            return
        
        try:
            self._export_to_excel(file_path)
            self.logger.info(f"📊 Test cases exported to Excel: {file_path}")
            messagebox.showinfo("Export Success", f"Test cases successfully exported to:\n{file_path}")
        except Exception as e:
            self.logger.error(f"❌ Excel export failed: {e}")
            messagebox.showerror("Export Error", f"Failed to export Excel:\n{e}")
    
    def _export_to_csv(self, file_path):
        """Export test cases to CSV format."""
        # Debug: Log quality report status
        self.logger.info(f"🔍 CSV Export Debug - Quality report available: {bool(self.quality_report)}")
        if self.quality_report:
            individual_scores = self.quality_report.get("individual_scores", [])
            self.logger.info(f"🔍 Individual scores count: {len(individual_scores)}")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            # Write header
            writer.writeheader()
            
            # Create quality score mapping
            quality_scores = {}
            if self.quality_report and "individual_scores" in self.quality_report:
                for score_info in self.quality_report["individual_scores"]:
                    test_id = score_info.get("test_id", "")
                    total_score = score_info.get("total_score", 0)
                    quality_scores[test_id] = total_score
                    self.logger.info(f"🔍 CSV Score mapping: {test_id} → {total_score:.1f}/10")
            
            # Write test cases
            for case in self.generated_cases:
                steps = case.get("steps", [])
                if isinstance(steps, list):
                    steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
                else:
                    steps_text = str(steps)
                
                test_id = case.get("id", "")
                quality_score = quality_scores.get(test_id, 0)
                quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
                
                writer.writerow({
                    'Test ID': test_id,
                    'Title': case.get("title", ""),
                    'Steps': steps_text,
                    'Expected Result': case.get("expected", ""),
                    'Priority': case.get("priority", "Medium"),
                    'Quality Score': quality_display
                })
            
            # Add metadata at the end
            writer.writerow({})  # Empty row
            writer.writerow({'Test ID': '--- METADATA ---'})
            writer.writerow({'Test ID': 'Generated On', 'Title': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            writer.writerow({'Test ID': 'Total Test Cases', 'Title': str(len(self.generated_cases))})
            
            if self.quality_report:
                overall_score = self.quality_report.get("overall_score", 0)
                writer.writerow({'Test ID': 'Overall Quality Score', 'Title': f"{overall_score:.1f}/10"})
    
    def _export_to_excel(self, file_path):
        """Export test cases to Excel format with formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Debug: Log quality report status
        self.logger.info(f"🔍 Excel Export Debug - Quality report available: {bool(self.quality_report)}")
        if self.quality_report:
            individual_scores = self.quality_report.get("individual_scores", [])
            self.logger.info(f"🔍 Individual scores count: {len(individual_scores)}")
            for score_info in individual_scores:
                test_id = score_info.get("test_id", "Unknown")
                total_score = score_info.get("total_score", 0)
                self.logger.info(f"🔍 Score mapping: {test_id} → {total_score:.1f}/10")
        
        wb = Workbook()
        
        # Test Cases Sheet
        ws_cases = wb.active
        ws_cases.title = "Test Cases"
        
        # Headers
        headers = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
        for col, header in enumerate(headers, 1):
            cell = ws_cases.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Create quality score mapping
        quality_scores = {}
        quality_details = {}
        if self.quality_report and "individual_scores" in self.quality_report:
            for score_info in self.quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                scores = score_info.get("scores", {})
                quality_scores[test_id] = total_score
                quality_details[test_id] = scores
        
        # Data rows
        for row, case in enumerate(self.generated_cases, 2):
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = "\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            # Debug: Log each quality score mapping
            self.logger.info(f"🔍 Excel row {row}: {test_id} → score={quality_score}, display='{quality_display}'")
            
            # Add data
            ws_cases.cell(row=row, column=1, value=test_id)
            ws_cases.cell(row=row, column=2, value=case.get("title", ""))
            
            steps_cell = ws_cases.cell(row=row, column=3, value=steps_text)
            steps_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws_cases.cell(row=row, column=4, value=case.get("expected", ""))
            
            priority_cell = ws_cases.cell(row=row, column=5, value=case.get("priority", "Medium"))
            
            quality_cell = ws_cases.cell(row=row, column=6, value=quality_display)
            self.logger.info(f"🔍 Set Excel cell F{row} = '{quality_display}'")
            
            # Color code quality scores
            if quality_score >= 8.0:
                quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif quality_score >= 6.0:
                quality_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
            elif quality_score > 0:
                quality_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Color code priority
            if case.get("priority", "").lower() == "high":
                priority_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            elif case.get("priority", "").lower() == "low":
                priority_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Adjust column widths
        ws_cases.column_dimensions['A'].width = 12  # Test ID
        ws_cases.column_dimensions['B'].width = 30  # Title
        ws_cases.column_dimensions['C'].width = 50  # Steps
        ws_cases.column_dimensions['D'].width = 30  # Expected
        ws_cases.column_dimensions['E'].width = 10  # Priority
        ws_cases.column_dimensions['F'].width = 15  # Quality
        
        # Quality Details Sheet
        if quality_details:
            ws_quality = wb.create_sheet("Quality Details")
            
            # Headers for quality details
            quality_headers = ['Test ID', 'Clarity', 'Completeness', 'Specificity', 'Testability', 'Coverage', 'Total Score']
            for col, header in enumerate(quality_headers, 1):
                cell = ws_quality.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Quality data
            for row, (test_id, scores) in enumerate(quality_details.items(), 2):
                ws_quality.cell(row=row, column=1, value=test_id)
                ws_quality.cell(row=row, column=2, value=f"{scores.get('clarity', 0):.1f}")
                ws_quality.cell(row=row, column=3, value=f"{scores.get('completeness', 0):.1f}")
                ws_quality.cell(row=row, column=4, value=f"{scores.get('specificity', 0):.1f}")
                ws_quality.cell(row=row, column=5, value=f"{scores.get('testability', 0):.1f}")
                ws_quality.cell(row=row, column=6, value=f"{scores.get('coverage', 0):.1f}")
                
                total_score = quality_scores.get(test_id, 0)
                total_cell = ws_quality.cell(row=row, column=7, value=f"{total_score:.1f}")
                
                # Color code total score
                if total_score >= 8.0:
                    total_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif total_score >= 6.0:
                    total_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
                elif total_score > 0:
                    total_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Adjust column widths
            for col in range(1, 8):
                ws_quality.column_dimensions[chr(64 + col)].width = 12
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        
        # Summary content
        summary_data = [
            ["Test Case Export Summary", ""],
            ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Test Cases", len(self.generated_cases)],
            ["", ""],
            ["Quality Assessment", ""],
        ]
        
        if self.quality_report:
            overall_score = self.quality_report.get("overall_score", 0)
            summary_data.extend([
                ["Overall Quality Score", f"{overall_score:.1f}/10"],
                ["", ""],
                ["Quality Distribution", ""],
            ])
            
            # Quality distribution
            if quality_scores:
                high_quality = sum(1 for score in quality_scores.values() if score >= 8.0)
                medium_quality = sum(1 for score in quality_scores.values() if 6.0 <= score < 8.0)
                low_quality = sum(1 for score in quality_scores.values() if score < 6.0)
                
                summary_data.extend([
                    ["High Quality (8.0+)", high_quality],
                    ["Medium Quality (6.0-7.9)", medium_quality],
                    ["Low Quality (<6.0)", low_quality],
                ])
        
        # Write summary data
        for row, (label, value) in enumerate(summary_data, 1):
            label_cell = ws_summary.cell(row=row, column=1, value=label)
            value_cell = ws_summary.cell(row=row, column=2, value=value)
            
            if label in ["Test Case Export Summary", "Quality Assessment", "Quality Distribution"]:
                label_cell.font = Font(bold=True, size=12)
            elif label == "":
                continue  # Skip empty rows
            else:
                label_cell.font = Font(bold=True)
        
        # Adjust summary column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        
        # Save workbook
        wb.save(file_path)


class LogHandler(logging.Handler):
    """Custom logging handler to display logs in the UI."""
    
    def __init__(self, app):
        super().__init__()
        self.app = app
    
    def emit(self, record):
        """Display log message in the UI."""
        try:
            msg = self.format(record)
            # Schedule UI update in main thread
            self.app.root.after(0, self._update_log, msg)
        except Exception:
            pass  # Ignore logging errors
    
    def _update_log(self, message):
        """Update log display (called from main thread)."""
        try:
            if hasattr(self.app, 'log_text'):
                self.app.log_text.insert(tk.END, message + "\n")
                self.app.log_text.see(tk.END)  # Auto-scroll to bottom
        except Exception:
            pass


def main():
    """Main entry point for the desktop application."""
    root = tk.Tk()
    
    # Apply modern theme for better UI appearance
    try:
        sv_ttk.set_theme("light")  # Use modern light theme
        print("✨ Applied modern theme successfully")
    except Exception as e:
        print(f"⚠️  Could not apply modern theme: {e}")
    
    app = TestCaseGeneratorApp(root)
    
    # Handle window closing
    def on_closing():
        if messagebox.askokcancel("Quit", "Do you want to quit the TestCase Generator?"):
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Start the GUI
    root.mainloop()


if __name__ == "__main__":
    main()