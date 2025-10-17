#!/usr/bin/env python3
"""
Test to reproduce the exact UI Excel export behavior
"""

import sys
from pathlib import Path
import json
from datetime import datetime

# Add the parent directory to sys.path  
sys.path.insert(0, str(Path(__file__).parent))

from src.core import chat, parse_json_safely, score_test_cases

def simulate_ui_excel_export():
    """Simulate exactly what happens in the UI when exporting to Excel"""
    print("🔍 Simulating UI Excel Export Process")
    print("=" * 50)
    
    # Step 1: Generate test cases (like UI does)
    print("📋 Step 1: Simulating test case generation...")
    
    # These are the kinds of cases the UI might generate
    generated_cases = [
        {
            "id": "TC-001",
            "title": "User login with valid credentials",
            "steps": [
                "Navigate to login page",
                "Enter valid username",
                "Enter valid password", 
                "Click login button"
            ],
            "expected": "User is logged in successfully",
            "priority": "High"
        },
        {
            "id": "TC-002", 
            "title": "User login with invalid credentials",
            "steps": [
                "Navigate to login page",
                "Enter invalid username or password",
                "Click login button"
            ],
            "expected": "Error message is displayed",
            "priority": "High"
        }
    ]
    
    requirement_text = "User login system with authentication functionality"
    
    print(f"✅ Generated {len(generated_cases)} test cases")
    
    # Step 2: Quality assessment (like UI does)
    print("\n📊 Step 2: Running quality assessment...")
    output_dir = Path("outputs/testcase_generated")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        quality_report = score_test_cases(generated_cases, requirement_text, output_dir)
        print("✅ Quality assessment completed")
        
        if quality_report:
            overall_score = quality_report.get("overall_score", 0)
            individual_scores = quality_report.get("individual_scores", [])
            print(f"   Overall Score: {overall_score:.1f}/10")
            print(f"   Individual Scores: {len(individual_scores)} cases")
        else:
            print("❌ No quality report generated")
            quality_report = {}
            
    except Exception as e:
        print(f"❌ Quality assessment failed: {e}")
        quality_report = {}
    
    # Step 3: Excel export (like UI _export_to_excel method)
    print("\n📊 Step 3: Excel export simulation...")
    
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = output_dir / f"ui_simulation_{timestamp}.xlsx"
        
        wb = Workbook()
        ws_cases = wb.active
        ws_cases.title = "Test Cases"
        
        # Headers (exactly like UI)
        headers = ['Test ID', 'Title', 'Steps', 'Expected Result', 'Priority', 'Quality Score']
        for col, header in enumerate(headers, 1):
            cell = ws_cases.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Create quality score mapping (exactly like UI)
        quality_scores = {}
        quality_details = {}
        if quality_report and "individual_scores" in quality_report:
            for score_info in quality_report["individual_scores"]:
                test_id = score_info.get("test_id", "")
                total_score = score_info.get("total_score", 0)
                scores = score_info.get("scores", {})
                quality_scores[test_id] = total_score
                quality_details[test_id] = scores
        
        print(f"   Quality scores mapping: {quality_scores}")
        print(f"   Quality details available: {len(quality_details)} cases")
        
        # Data rows (exactly like UI)
        for row, case in enumerate(generated_cases, 2):
            steps = case.get("steps", [])
            if isinstance(steps, list):
                steps_text = "\\n".join(f"{i+1}. {step}" for i, step in enumerate(steps))
            else:
                steps_text = str(steps)
            
            test_id = case.get("id", "")
            quality_score = quality_scores.get(test_id, 0)
            quality_display = f"{quality_score:.1f}/10" if quality_score > 0 else "N/A"
            
            print(f"   Processing {test_id}: quality_score={quality_score}, display='{quality_display}'")
            
            # Add data (exactly like UI)
            ws_cases.cell(row=row, column=1, value=test_id)
            ws_cases.cell(row=row, column=2, value=case.get("title", ""))
            
            steps_cell = ws_cases.cell(row=row, column=3, value=steps_text)
            steps_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            ws_cases.cell(row=row, column=4, value=case.get("expected", ""))
            
            priority_cell = ws_cases.cell(row=row, column=5, value=case.get("priority", "Medium"))
            
            quality_cell = ws_cases.cell(row=row, column=6, value=quality_display)
            print(f"   Set quality cell F{row} = '{quality_display}'")
            
            # Color code quality scores (exactly like UI)
            if quality_score >= 8.0:
                quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                print(f"   Applied GREEN color to {test_id}")
            elif quality_score >= 6.0:
                quality_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
                print(f"   Applied YELLOW color to {test_id}")
            elif quality_score > 0:
                quality_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                print(f"   Applied PINK color to {test_id}")
            else:
                print(f"   No color applied to {test_id} (score: {quality_score})")
        
        # Adjust column widths (exactly like UI)
        ws_cases.column_dimensions['A'].width = 12
        ws_cases.column_dimensions['B'].width = 30
        ws_cases.column_dimensions['C'].width = 50
        ws_cases.column_dimensions['D'].width = 30
        ws_cases.column_dimensions['E'].width = 10
        ws_cases.column_dimensions['F'].width = 15
        
        # Save workbook
        wb.save(excel_file)
        print(f"✅ Excel file saved: {excel_file}")
        
        # Verify by reading back (critical step)
        print("\\n🔍 Step 4: Verifying Excel content...")
        wb_verify = openpyxl.load_workbook(excel_file)
        ws_verify = wb_verify.active
        
        print(f"   Workbook title: {ws_verify.title}")
        print(f"   Dimensions: {ws_verify.max_row} rows x {ws_verify.max_column} columns")
        
        # Check headers
        print("   Headers:")
        for col in range(1, 7):
            header = ws_verify.cell(row=1, column=col).value
            print(f"     Column {col}: '{header}'")
        
        # Check data with quality scores
        print("   Data rows with quality scores:")
        for row in range(2, ws_verify.max_row + 1):
            test_id = ws_verify.cell(row=row, column=1).value
            title = ws_verify.cell(row=row, column=2).value
            quality_value = ws_verify.cell(row=row, column=6).value
            print(f"     Row {row}: {test_id} | {title[:20]}... | Quality: '{quality_value}'")
            
            # Check if quality score is missing
            if quality_value is None or quality_value == "" or quality_value == "N/A":
                print(f"     ❌ ISSUE: No quality score for {test_id}")
            else:
                print(f"     ✅ Quality score found: {quality_value}")
        
        wb_verify.close()
        
        print(f"\\n📊 Excel Export Simulation Results:")
        print(f"   ✅ File created: {excel_file}")
        print(f"   ✅ Size: {excel_file.stat().st_size} bytes")
        print(f"   ✅ Quality scores: {len(quality_scores)} mapped")
        print(f"   ✅ Data rows: {len(generated_cases)} test cases")
        
        return True
        
    except ImportError:
        print("❌ openpyxl not available")
        return False
    except Exception as e:
        print(f"❌ Excel export simulation failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    success = simulate_ui_excel_export()
    if success:
        print("\\n✅ UI Excel export simulation completed successfully!")
        print("\\n🔍 Analysis:")
        print("   If quality scores are visible in this test but not in the UI,")
        print("   the issue might be:")
        print("   1. Quality assessment not running properly in UI")
        print("   2. Quality report not being passed to export function")
        print("   3. UI state management issue with quality_report variable")
        print("\\n📝 Next step: Check the UI application's quality_report variable")
    else:
        print("\\n❌ Simulation failed - check errors above")

if __name__ == "__main__":
    main()